import os
import time
import json
import logging
import random
import re
from datetime import datetime
import subprocess
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options 
import undetected_chromedriver as uc

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

class SubstackHandler:
    def __init__(self, driver, email, password):
        self.driver = driver
        self.email = email
        self.password = password
        self.wait = WebDriverWait(driver, 30)
        self.screenshots_dir = "screenshots"
        self.blog_file = "blog.txt"
        os.makedirs(self.screenshots_dir, exist_ok=True)
    
    def _resolve_image_path(self, path_str):
        base_dir = os.path.dirname(os.path.abspath(self.blog_file)) if getattr(self, "blog_file", None) else os.getcwd()
        s = (path_str or "").strip()
        if not s:
            return None
        candidates = []
        if ('/' in s) or ('\\' in s):
            candidates.extend([
                s,
                os.path.abspath(s) if os.path.isabs(s) else os.path.abspath(os.path.join(base_dir, s)),
                os.path.join(base_dir, s.replace('../', '').replace('..\\', '')),
                s.replace('../', '').replace('..\\', ''),
            ])
        else:
            candidates.extend([
                os.path.abspath(os.path.join(base_dir, '..', 'Image', s)),
                os.path.join(base_dir, 'Image', s),
                os.path.join(base_dir, s),
                os.path.join(base_dir, 'Ima', s),
                os.path.abspath(os.path.join(base_dir, '..', 'Ima', s)),
            ])
        for p in candidates:
            try:
                ap = os.path.abspath(p)
                if os.path.isfile(ap):
                    return ap
            except Exception:
                continue
        return None
    
    def take_screenshot(self, step_name):
        """Take debug screenshot"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.screenshots_dir}/{step_name}_{timestamp}.png"
            self.driver.save_screenshot(filename)
            logger.info(f"📸 Screenshot saved: {filename}")
        except Exception as e:
            logger.error(f"Failed to take screenshot: {e}")

    def human_type(self, element, text, min_delay=0.005, max_delay=0.02):
        """Robust typing: Sends text in chunks or all at once"""
        try:
            # Try sending all at once first as it is most reliable
            element.send_keys(text)
            time.sleep(0.1)
        except Exception as e:
            logger.warning(f"Error in fast typing, falling back to char-by-char: {e}")
            # Fallback to char by char if bulk send fails
            for ch in text:
                try:
                    element.send_keys(ch)
                    time.sleep(random.uniform(min_delay, max_delay))
                except Exception as e:
                    logger.warning(f"Error typing char '{ch}': {e}")

    def ensure_editor_focus(self, element):
        """Ensure editor is in view and focused"""
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", element)
            time.sleep(0.1)
        except:
            pass

    def type_line_with_links(self, element, line):
        """Type a line handling Markdown links [text](url)"""
        try:
            self.ensure_editor_focus(element)
            
            # Find all links in the line
            cursor = 0
            while cursor < len(line):
                # Find next link starting from cursor
                match = re.search(r'\[(.*?)\]\((.*?)\)', line[cursor:])
                
                if match:
                    # Calculate absolute positions
                    start_idx = cursor + match.start()
                    end_idx = cursor + match.end()
                    
                    # Type text before link
                    prefix = line[cursor:start_idx]
                    if prefix:
                        self.human_type(element, prefix)
                        
                    # Link Text and URL
                    link_text = match.group(1)
                    url = match.group(2)
                    
                    # Type link text
                    self.human_type(element, link_text)
                    
                    # Select link text (backwards)
                    actions = ActionChains(self.driver)
                    actions.key_down(Keys.SHIFT)
                    for _ in range(len(link_text)):
                        actions.send_keys(Keys.ARROW_LEFT)
                    actions.key_up(Keys.SHIFT)
                    actions.perform()
                    time.sleep(0.2)
                    
                    # Open Link Dialog (Ctrl+K)
                    actions = ActionChains(self.driver)
                    actions.key_down(Keys.CONTROL).send_keys('k').key_up(Keys.CONTROL)
                    actions.perform()
                    time.sleep(1)
                    
                    # Type URL
                    actions = ActionChains(self.driver)
                    actions.send_keys(url)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(0.5)
                    
                    # Move cursor to end of line (Right arrow)
                    element.send_keys(Keys.ARROW_RIGHT)
                    
                    cursor = end_idx
                else:
                    # No more links, type rest
                    remaining = line[cursor:]
                    if remaining:
                        self.human_type(element, remaining)
                    break
            return True
        except Exception as e:
            logger.error(f"Error typing line with links: {e}")
            # Fallback: type raw line
            try:
                self.human_type(element, line)
            except:
                pass
            return False

    def simulate_mouse_move(self, element):
        """Simulate moving mouse to element"""
        try:
            action = ActionChains(self.driver)
            action.move_to_element(element)
            action.perform()
            time.sleep(random.uniform(0.2, 0.5))
        except:
            pass
        
    def get_clipboard_text(self):
        try:
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", "Get-Clipboard"],
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode == 0:
                text = result.stdout.strip()
                return text if text else None
            return None
        except:
            return None
    
    def save_url_to_file(self, url):
        try:
            path = "url.txt"
            lines = []
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
            idx = None
            for i, line in enumerate(lines):
                if not line.strip():
                    idx = i
                    break
            if idx is not None:
                lines[idx] = url.strip() + "\n"
                with open(path, "w", encoding="utf-8") as f:
                    f.writelines(lines)
            else:
                with open(path, "a", encoding="utf-8") as f:
                    f.write(url.strip() + "\n")
            return True
        except:
            return False
        
    def click_password_link(self):
        try:
            selector = "//*[self::a or self::button or self::span][contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in with password') or contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in with password')]"
            elems = self.driver.find_elements(By.XPATH, selector)
            if not elems:
                return False
            target = None
            for el in elems:
                if el.is_displayed():
                    target = el
                    break
            if not target:
                target = elems[0]
            try:
                self.simulate_mouse_move(target)
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
                self.driver.execute_script("arguments[0].click();", target)
            except:
                try:
                    ActionChains(self.driver).move_to_element(target).click().perform()
                except:
                    try:
                        target.click()
                    except:
                        return False
            time.sleep(2)
            return True
        except:
            return False
    
    def open_create_post(self):
        try:
            self.driver.get("https://substack.com/home")
            time.sleep(3)
            create_btn = None
            candidates = self.driver.find_elements(By.XPATH, "//button[contains(.,'Create')] | //span[contains(.,'Create')]/ancestor::*[self::button or self::a]")
            for el in candidates:
                if el.is_displayed():
                    create_btn = el
                    break
            if not create_btn and candidates:
                create_btn = candidates[0]
            if not create_btn:
                return False
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", create_btn)
                self.simulate_mouse_move(create_btn)
                self.driver.execute_script("arguments[0].click();", create_btn)
            except:
                create_btn.click()
            time.sleep(1)
            post_item = None
            items = self.driver.find_elements(By.XPATH, "//*[self::a or self::button][contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'post')]")
            for el in items:
                if el.is_displayed():
                    post_item = el
                    break
            if not post_item and items:
                post_item = items[0]
            if not post_item:
                return False
            try:
                self.simulate_mouse_move(post_item)
                self.driver.execute_script("arguments[0].click();", post_item)
            except:
                post_item.click()
            time.sleep(4)
            return True
        except:
            return False
    
    def extract_and_save_live_link(self):
        try:
            logger.info("🕵️ Looking for live link...")
            time.sleep(3)
            
            # Wait for "Share your link" section
            try:
                self.wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Share your link')]")))
            except:
                logger.warning("⚠️ 'Share your link' text not found, continuing anyway...")

            # 1. Click "Copy link" button
            try:
                copy_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Copy link')]")))
                self.simulate_mouse_move(copy_btn)
                copy_btn.click()
                logger.info("✅ Clicked 'Copy link' button")
                time.sleep(1)
            except Exception as e:
                logger.warning(f"⚠️ Could not click 'Copy link' button: {e}")

            clipboard_url = self.get_clipboard_text()
            if clipboard_url and clipboard_url.startswith("http"):
                self.save_url_to_file(clipboard_url)
                logger.info(f"🔗 Clipboard URL saved: {clipboard_url}")
                self.take_screenshot("07_live_link_saved")
                return True

            # 2. Extract URL from input field
            url = None
            try:
                # Look for input with http value
                inputs = self.driver.find_elements(By.XPATH, "//input[contains(@value, 'http')]")
                for inp in inputs:
                    if inp.is_displayed():
                        val = inp.get_attribute("value")
                        if "substack.com" in val:
                            url = val
                            break
                
                if not url:
                    # Fallback: Look for any element with http text that looks like a url
                    candidates = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'https://') and contains(text(), 'substack.com')]")
                    for cand in candidates:
                        if cand.is_displayed():
                            url = cand.text.strip()
                            break
            except Exception as e:
                logger.error(f"Error finding URL input: {e}")

            # 3. Save to file
            if url:
                logger.info(f"🔗 Found URL: {url}")
                self.save_url_to_file(url)
                logger.info("💾 Saved URL to url.txt")
                self.take_screenshot("07_live_link_saved")
                return True
            else:
                logger.error("❌ Could not find URL in the page.")
                self.take_screenshot("error_link_not_found")
                return False

        except Exception as e:
            logger.error(f"❌ Error in extract_and_save_live_link: {e}")
            return False

    def login(self):
        """Login to Substack using Password"""
        try:
            logger.info("🚀 Navigating to Substack Login...")
            self.driver.get("https://substack.com/sign-in")
            time.sleep(random.uniform(3, 5))
            self.take_screenshot("01_login_page")

            logger.info("Entering Email...")
            email_input = self.wait.until(EC.presence_of_element_located((By.NAME, "email")))
            self.human_type(email_input, self.email)
            time.sleep(1)
            try:
                email_input.send_keys(Keys.TAB)
                self.driver.execute_script("arguments[0].blur();", email_input)
            except:
                pass

            logger.info("Opening password form...")
            try:
                if not self.click_password_link():
                    self.take_screenshot("debug_password_link_not_found")
                    return False
            except:
                self.take_screenshot("debug_password_link_click_failed")
                return False

            logger.info("Entering Password...")
            try:
                pass_input = self.wait.until(EC.presence_of_element_located((By.NAME, "password")))
            except:
                self.take_screenshot("debug_no_password_field")
                return False

            self.human_type(pass_input, self.password)
            time.sleep(1)

            logger.info("Submitting...")
            try:
                submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit' or contains(text(),'Continue')]")
            except:
                submit_btn = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            self.simulate_mouse_move(submit_btn)
            submit_btn.click()

            time.sleep(5)
            if "challenge" in self.driver.current_url:
                time.sleep(15)

            self.wait.until(lambda d: "sign-in" not in d.current_url)
            self.take_screenshot("02_logged_in")
            return True

        except Exception as e:
            logger.error(f"❌ Login Failed: {e}")
            self.take_screenshot("error_login")
            return False

    def copy_image_to_clipboard(self, image_path):
        """Copy image file to clipboard using PowerShell"""
        try:
            abs_path = os.path.abspath(image_path)
            cmd = f'''
            Add-Type -AssemblyName System.Windows.Forms
            $files = [System.Collections.Specialized.StringCollection]::new()
            $files.Add("{abs_path}")
            [System.Windows.Forms.Clipboard]::SetFileDropList($files)
            '''
            subprocess.run(["powershell", "-command", cmd], timeout=5)
            return True
        except Exception as e:
            logger.error(f"Failed to copy to clipboard: {e}")
            return False

    def upload_image(self, image_path):
        """Upload image to Substack editor using Clipboard Paste"""
        try:
            logger.info(f"📸 Uploading image: {image_path}")
            if not os.path.exists(image_path):
                logger.error(f"Image file not found: {image_path}")
                return False

            # Method 1: Clipboard Paste (Best for rich text editors)
            if self.copy_image_to_clipboard(image_path):
                time.sleep(1)
                # Paste (Ctrl+V)
                actions = ActionChains(self.driver)
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL)
                actions.perform()
                logger.info("✅ Pasted image from clipboard")
                time.sleep(5) 
                return True
                
            # Method 2: Fallback to file input
            file_inputs = self.driver.find_elements(By.XPATH, "//input[@type='file']")
            if file_inputs:
                file_input = file_inputs[-1] 
                file_input.send_keys(image_path)
                time.sleep(5) 
                return True
            
            logger.warning("Could not find file input for image upload")
            return False
        except Exception as e:
            logger.error(f"Error uploading image: {e}")
            return False

    def create_and_publish_post(self, title, content):
        """Create a new post and publish it"""
        try:
            logger.info("📝 Starting New Post...")
            
            if not self.open_create_post():
                try:
                    self.driver.get("https://substack.com/publish/post")
                    time.sleep(5)
                except:
                    return False
            self.take_screenshot("03_editor_loaded")

            # 2. Enter Title
            logger.info("Entering Title...")
            try:
                # Substack title is usually a textarea with placeholder "Title"
                title_area = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "textarea[placeholder='Title']")))
                title_area.click()
                self.human_type(title_area, title)
            except Exception as e:
                logger.error("Could not find Title field")
                return False

            time.sleep(1)

            # 3. Enter Content
            logger.info("Entering Content...")
            try:
                # Substack editor is often a contenteditable div
                # We look for the main editor body
                editor_div = self.driver.find_element(By.CSS_SELECTOR, "div.pencraft.pc-editor, div[contenteditable='true']")
                editor_div.click()
                time.sleep(0.5)
                
                # Type content (handling newlines and links)
                # content is now a list of lines
                for i, line in enumerate(content):
                    logger.info(f"Typing line {i+1}/{len(content)}")
                    
                    # Check for image path pattern: {path}
                    # Matches {C:\Path\To\Image.jpg}
                    img_match = re.match(r'^\s*\{(.*?)\}\s*$', line.strip())
                    if img_match:
                        image_path = img_match.group(1).strip()
                        logger.info(f"🖼️ Found image path: {image_path}")
                        
                        resolved = self._resolve_image_path(image_path)
                        if resolved and self.upload_image(resolved):
                             # Ensure focus is back in editor and maybe add a newline
                            try:
                                # Wait for image to render
                                time.sleep(2)
                                
                                # Move out of caption/image context
                                # 1. Escape to deselect image/caption
                                # 2. Arrow Down to move cursor below image
                                # 3. Enter to start new paragraph
                                actions = ActionChains(self.driver)
                                actions.send_keys(Keys.ESCAPE)
                                actions.pause(0.5)
                                actions.send_keys(Keys.ARROW_DOWN)
                                actions.pause(0.5)
                                actions.send_keys(Keys.ENTER)
                                actions.perform()
                                
                                # Double check by sending another Enter to editor
                                time.sleep(0.5)
                                editor_div.send_keys(Keys.ENTER)
                            except Exception as e:
                                logger.warning(f"Error navigating after image: {e}")
                        else:
                            logger.error(f"❌ Failed to upload image: {image_path}")
                            try:
                                self.type_line_with_links(editor_div, line)
                                editor_div.send_keys(Keys.ENTER)
                            except:
                                pass
                        continue
                    
                    if not line.strip():
                        editor_div.send_keys(Keys.ENTER)
                        time.sleep(0.1)
                        continue

                    # Use the new helper to type line with links
                    self.type_line_with_links(editor_div, line)
                        
                    editor_div.send_keys(Keys.ENTER)
                    time.sleep(0.1)
                    
            except Exception as e:
                logger.error(f"Could not find Content editor or error typing: {e}")
                return False

            time.sleep(3)
            self.take_screenshot("04_content_filled")

            logger.info("Clicking Continue...")
            try:
                continue_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Continue')]")))
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", continue_btn)
                self.simulate_mouse_move(continue_btn)
                try:
                    self.driver.execute_script("arguments[0].click();", continue_btn)
                except:
                    continue_btn.click()
            except:
                self.take_screenshot("error_continue_not_found")
                return False

            time.sleep(3)
            self.take_screenshot("05_publish_settings")

            logger.info("Finalizing Publication...")
            
            publish_xpath = "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'send to everyone now') or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'send to everyone') or contains(., 'Publish')]"
            
            try:
                self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='dialog'] | //div[contains(@class,'modal')]")))
                final_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, publish_xpath)))
                
                self.simulate_mouse_move(final_btn)
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", final_btn)
                    self.driver.execute_script("arguments[0].click();", final_btn)
                except:
                    final_btn.click()
                logger.info("🎉 Publish button clicked!")
                
                try:
                    time.sleep(1)
                    pwb_btn = self.wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "//*[self::button or self::a][contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'publish without buttons')]")
                        )
                    )
                    self.simulate_mouse_move(pwb_btn)
                    try:
                        self.driver.execute_script("arguments[0].click();", pwb_btn)
                    except:
                        pwb_btn.click()
                    time.sleep(2)
                    logger.info("✅ Publish without buttons confirmed")
                except:
                    pass
                
            except:
                logger.warning("Could not click final publish button. Check screenshot.")
                return False

            time.sleep(5)
            self.take_screenshot("06_published")
            
            if "post" in self.driver.current_url or "publish" not in self.driver.current_url:
                logger.info("✅ Post Published Successfully!")
                try:
                    self.extract_and_save_live_link()
                except:
                    pass
                return True
            
            return False

        except Exception as e:
            logger.error(f"❌ Error in creating post: {e}")
            self.take_screenshot("error_post_creation")
            return False

def read_blog_data():
    """Read data from blog.txt"""
    try:
        if not os.path.exists("blog.txt"):
            # Create dummy file if not exists
            with open("blog.txt", "w", encoding="utf-8") as f:
                f.write("My Substack Automation Journey\n\nThis is a test post created by Python script.")
            return "My Substack Automation Journey", ["This is a test post created by Python script."]
            
        with open("blog.txt", "r", encoding="utf-8") as f:
            lines = f.readlines()
            
        title = lines[0].strip() if lines else "Untitled Post"
        # Return content as list of lines to preserve structure
        content_lines = [line.rstrip() for line in lines[1:]]
        return title, content_lines
    except:
        return "Test Title", ["Test Content"]

def setup_driver():
    """Setup Undetected Chrome Driver"""
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    
    # Use a real user profile if possible to avoid captcha
    # options.add_argument(f"--user-data-dir={os.getenv('LOCALAPPDATA')}\\Google\\Chrome\\User Data")
    
    try:
        driver = uc.Chrome(options=options, headless=False)
    except:
        driver = uc.Chrome(headless=False)
        
    return driver

def read_credentials_from_excel(file_path="credentials.xlsx"):
    """Read Substack credentials from Excel"""
    try:
        if not os.path.exists(file_path):
            logger.warning(f"File not found: {file_path}")
            return None

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Simple header mapping (assuming first row is header)
        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[str(val).lower().strip()] = col
                
        # Helper to get col index by keywords
        def get_col_idx(keywords):
            for k, v in headers.items():
                if any(kw in k for kw in keywords):
                    return v
            return None
            
        website_col = get_col_idx(['website', 'url', 'site'])
        email_col = get_col_idx(['email', 'mail'])
        username_col = get_col_idx(['username', 'user'])
        password_col = get_col_idx(['password', 'pass', 'pwd'])
        
        # Search for substack
        for row in range(2, sheet.max_row + 1):
            website = str(sheet.cell(row=row, column=website_col).value or "").strip() if website_col else ""
            if "substack.com" in website.lower():
                email = str(sheet.cell(row=row, column=email_col).value or "").strip() if email_col else ""
                username = str(sheet.cell(row=row, column=username_col).value or "").strip() if username_col else ""
                password = str(sheet.cell(row=row, column=password_col).value or "").strip() if password_col else ""
                
                if email and password:
                    return {
                        "website": website,
                        "email": email,
                        "username": username,
                        "password": password
                    }
        return None
    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        return None

def run(driver, website, email, username, password):
    """Entry point for script.py to run Substack automation using credentials.xlsx"""
    try:
        logger.info(f"Starting Substack run for {email}")
        bot = SubstackHandler(driver, email, password)
        
        # Login
        if not bot.login():
            logger.error("Login failed")
            return False
            
        # Read content
        title, content = read_blog_data()
        
        # Publish
        if bot.create_and_publish_post(title, content):
            logger.info("Substack post published successfully")
            return True
        else:
            logger.error("Failed to publish post")
            return False
            
    except Exception as e:
        logger.error(f"Error in Substack run: {e}")
        return False

def main():
    print("="*60)
    print("SUBSTACK AUTOMATION BOT")
    print("="*60)

    # 1. Credentials
    email = ""
    password = ""
    
    # Try to load from Excel
    creds = read_credentials_from_excel()
    if creds:
        print(f"✅ Loaded credentials for {creds['website']} from credentials.xlsx")
        email = creds['email']
        password = creds['password']
    else:
        print("⚠️ credentials.xlsx not found or no Substack row found.")
        print("Using hardcoded fallback (please update script if needed).")
        # Fallback
        email = "kavikr543@gmail.com"
        password = "YOUR_SUBSTACK_PASSWORD_HERE" 

    if password == "YOUR_SUBSTACK_PASSWORD_HERE" and not creds:
        print("❌ Error: Please update the 'password' variable in the script or add to credentials.xlsx")
        return

    # 2. Read Blog Content
    title, content = read_blog_data()
    print(f"📝 Post Title: {title}")

    # 3. Start Driver
    driver = None
    try:
        driver = setup_driver()
        bot = SubstackHandler(driver, email, password)

        # 4. Login
        if bot.login():
            # 5. Publish
            bot.create_and_publish_post(title, content)
        
        print("\nBrowser remaining open for inspection (60s)...")
        time.sleep(60)

    except Exception as e:
        print(f"Critical Error: {e}")
    finally:
        if driver:
            driver.quit()

if __name__ == "__main__":
    main()
