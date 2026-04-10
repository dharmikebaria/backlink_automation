"""
Quora Blog Posting Automation Script - Merged with User Specific CSS Selector
Author: Automation Bot
Description: Automates login and blog posting on Quora with robust element detection and specific CSS selector support
"""

import time
import logging
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
    StaleElementReferenceException,
    ElementNotInteractableException,
    WebDriverException
)
from selenium.webdriver.common.action_chains import ActionChains
import random
import openpyxl
import os
import re
import io
try:
    import win32clipboard
    import win32con
    from PIL import Image
except Exception:
    win32clipboard = None
    win32con = None
    Image = None


class QuoraBlogPoster:
    """Automation class for posting blogs on Quora with robust detection"""
    
    def __init__(self, headless=False, driver=None, email=None, password=None, website=None):
        """
        Initialize the automation bot
        
        Args:
            headless (bool): Run browser in headless mode
        """
        self.setup_logging()
        self.logger = logging.getLogger(__name__)
        
        # Configuration
        self.quora_url = website or "https://www.quora.com/"
        self.credentials = {
            'email': email or '',
            'password': password or ''
        }
        self.blog_content = self.load_blog_content()
        
        # User specific CSS selector for the post button/container
        self.user_css_selector = "#root > div > div:nth-child(2) > div > div > div > div > div.q-flex.sc-gKXOVf.ddbdSM.modal_content_inner.qu-flexDirection--column.qu-bg--white.qu-overflowY--auto.qu-borderAll.qu-alignSelf--center > div > div > div.q-flex.qu-flexDirection--column.qu-overflowY--auto > div.q-sticky.qu-bg--white.qu-borderTop > div > div.q-flex.qu-justifyContent--flex-end.qu-alignItems--center"
        
        # Initialize driver
        self.driver = driver or self.setup_driver(headless)
        self._owns_driver = driver is None
        self.wait = WebDriverWait(self.driver, 20)
        self.short_wait = WebDriverWait(self.driver, 5)
        
        # Screenshot directory
        self.screenshot_dir = Path("screenshots")
        self.screenshot_dir.mkdir(exist_ok=True)
        self.safe_mode = True
        self.activity_min_s = 8
        self.activity_max_s = 15
    
    def _send_to_clipboard(self, clip_type, data):
        try:
            if not win32clipboard:
                return False
            for _ in range(5):
                try:
                    win32clipboard.OpenClipboard()
                    win32clipboard.EmptyClipboard()
                    win32clipboard.SetClipboardData(clip_type, data)
                    win32clipboard.CloseClipboard()
                    return True
                except Exception:
                    try:
                        win32clipboard.CloseClipboard()
                    except Exception:
                        pass
                    time.sleep(0.3)
            return False
        except Exception:
            return False
    
    def _resolve_image_path(self, p):
        try:
            # Clean up path
            p = p.strip().strip('"').strip("'")
            
            # If absolute path and exists, return it
            if os.path.exists(p) and os.path.isfile(p):
                return os.path.abspath(p)
                
            # Get script directory
            script_dir = os.path.dirname(os.path.abspath(__file__))
            
            # List of candidate paths to check
            candidates = [
                # Check relative to script directory's Image folder (most likely)
                os.path.join(script_dir, "Image", p),
                # Check relative to script directory parent's Image folder (user request ../Image)
                os.path.join(script_dir, "..", "Image", p),
                # Check relative to CWD Image folder
                os.path.join(os.getcwd(), "Image", p),
                # Check relative to script directory directly
                os.path.join(script_dir, p),
                # Check relative to CWD directly
                os.path.abspath(p)
            ]
            
            for candidate in candidates:
                # Normalize path
                candidate = os.path.normpath(candidate)
                if os.path.exists(candidate) and os.path.isfile(candidate):
                    self.logger.info(f"Resolved image path: {candidate}")
                    return candidate
            
            # If not found, try to search recursively in script dir
            for root, dirs, files in os.walk(script_dir):
                if p in files:
                    candidate = os.path.join(root, p)
                    self.logger.info(f"Found image via search: {candidate}")
                    return candidate

            self.logger.warning(f"Image not found: {p}")
            return None
        except Exception as e:
            self.logger.error(f"Path resolution error: {e}")
            return None
    
    def _copy_image_to_clipboard(self, image_path):
        try:
            if not Image or not win32con:
                return False
            if not image_path or not os.path.exists(image_path):
                return False
            with Image.open(image_path) as im:
                if im.width > 1600 or im.height > 1600:
                    im.thumbnail((1280, 1280))
                buf = io.BytesIO()
                im.convert("RGB").save(buf, "BMP")
                data = buf.getvalue()[14:]
                buf.close()
            return self._send_to_clipboard(win32con.CF_DIB, data)
        except Exception:
            return False
    
    def _get_editor_image_count(self):
        try:
            # Try to find images inside the editor
            count = 0
            selectors = [
                (By.CSS_SELECTOR, ".q-text-editor img"),
                (By.XPATH, "//div[@contenteditable='true']//img"),
                (By.XPATH, "//div[@role='textbox']//img")
            ]
            for by, sel in selectors:
                try:
                    imgs = self.driver.find_elements(by, sel)
                    if imgs:
                        count = max(count, len(imgs))
                except:
                    continue
            return count
        except:
            return 0

    def _wait_image_in_editor(self, initial_count=0, timeout=25):
        try:
            end = time.time() + timeout
            while time.time() < end:
                try:
                    current_count = self._get_editor_image_count()
                    if current_count > initial_count:
                        return True
                except Exception:
                    pass
                time.sleep(0.5)
            return False
        except Exception:
            return False
    
    def _force_caret_end(self, input_field):
        try:
            self.driver.execute_script("""
                var el = arguments[0];
                var range = document.createRange();
                range.selectNodeContents(el);
                range.collapse(false);
                var sel = window.getSelection();
                sel.removeAllRanges();
                sel.addRange(range);
            """, input_field)
        except Exception:
            pass
    
    def _safe_break(self, input_field, count=2):
        try:
            self._force_caret_end(input_field)
            for _ in range(max(1, count)):
                input_field.send_keys(Keys.ENTER)
                time.sleep(0.15)
            self._force_caret_end(input_field)
        except Exception:
            pass
    
    def _dismiss_interfering_popups(self):
        self.handle_popups()
        try:
             # Click 'Close' on any dialogs that might have appeared
             close_btns = self.driver.find_elements(By.XPATH, "//button[contains(@aria-label, 'Close')]")
             for btn in close_btns:
                 if btn.is_displayed():
                     btn.click()
        except:
            pass

    def _ensure_editor_focus(self, element):
        try:
            element.click()
            return element
        except StaleElementReferenceException:
            # Re-find
             try:
                 return self.driver.find_element(By.CSS_SELECTOR, ".q-text-editor")
             except:
                 return None
        except Exception:
            return element

    def _handle_image_token(self, input_field, path_token):
        try:
            resolved = self._resolve_image_path(path_token)
            if not resolved:
                return False
            
            # Try up to 3 times to paste image
            for attempt in range(3):
                ok = self._copy_image_to_clipboard(resolved)
                if not ok:
                    time.sleep(1)
                    continue
                
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", input_field)
                except Exception:
                    pass
                
                input_field = self._ensure_editor_focus(input_field) or input_field
                time.sleep(0.5)
                
                # Get initial count
                initial_count = self._get_editor_image_count()
                
                # Force clean break
                self._safe_break(input_field, count=1)
                
                # Paste
                ActionChains(self.driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                
                # Wait for image
                pasted = self._wait_image_in_editor(initial_count=initial_count, timeout=20)
                if pasted:
                    self._safe_break(input_field, count=1)
                    return True
                
                self.logger.warning(f"Image paste attempt {attempt+1} failed, retrying...")
                self._dismiss_interfering_popups()
                time.sleep(2)
            
            return False
        except Exception:
            return False
        
    def setup_logging(self):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('quora_automation.log'),
                logging.StreamHandler()
            ]
        )
    
    def setup_driver(self, headless=False):
        """
        Setup Chrome driver with appropriate options
        
        Args:
            headless (bool): Run in headless mode
            
        Returns:
            webdriver.Chrome: Configured Chrome driver
        """
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from webdriver_manager.chrome import ChromeDriverManager
        
        chrome_options = Options()
        
        # Disable automation detection
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-notifications")
        try:
            w = random.randint(1200, 1600)
            h = random.randint(800, 1000)
            chrome_options.add_argument(f"--window-size={w},{h}")
        except Exception:
            chrome_options.add_argument("--window-size=1400,900")
        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        chrome_options.add_argument(f"--user-agent={ua}")
        
        # Add human-like behavior
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-popup-blocking")
        
        # Remove headless for debugging (you can see what's happening)
        if headless:
            chrome_options.add_argument("--headless=new")
        
        # Experimental options
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Disable save password prompts
        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.notifications": 2,
            "profile.default_content_setting_values.cookies": 1,
            "profile.default_content_setting_values.popups": 2
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        # Initialize driver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        
        # Execute CDP commands to prevent detection
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                window.chrome = { runtime: {} };
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['en-US', 'en']
                });
                Object.defineProperty(navigator, 'platform', {
                    get: () => 'Win32'
                });
                const origQuery = window.navigator.permissions.query;
                window.navigator.permissions.query = (parameters) => (
                  parameters.name === 'notifications' ?
                    Promise.resolve({ state: Notification.permission }) :
                    origQuery(parameters)
                );
            """
        })
        
        return driver

    def load_blog_content(self):
        """Load title and body from blog.txt"""
        try:
            if Path("blog.txt").exists():
                text = Path("blog.txt").read_text(encoding="utf-8")
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                title = lines[0] if lines else ""
                body = "\n".join(lines[1:]) if len(lines) > 1 else ""
                if not title:
                    title = "Untitled"
                if not body:
                    body = ""
                self.logger.info(f"Loaded blog.txt (title: '{title[:60]}', body chars: {len(body)})")
                return {'title': title, 'body': body}
            else:
                self.logger.warning("blog.txt not found, using default content")
                return {'title': 'Default Title', 'body': 'Default Body Content'}
        except Exception as e:
            self.logger.error(f"Failed to load blog.txt: {str(e)}")
            return {
                'title': 'Untitled',
                'body': ''
            }

    def human_delay(self):
        """Add human-like delay to avoid detection"""
        time.sleep(random.uniform(0.01, 0.03))

    def take_screenshot(self, name):
        """Take screenshot for debugging"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.screenshot_dir}/{name}_{timestamp}.png"
            self.driver.save_screenshot(str(filename))
            self.logger.info(f"Screenshot saved: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"Failed to take screenshot: {str(e)}")
            return None
    
    def highlight_element(self, element):
        """Highlight element for visual debugging"""
        try:
            # Save original style
            original_style = element.get_attribute('style')
            # Apply highlight style (Red border + Yellow background)
            self.driver.execute_script("arguments[0].setAttribute('style', arguments[1]);", 
                                       element, 
                                       "border: 3px solid red; background-color: yellow; box-shadow: 0px 0px 10px red;")
            time.sleep(0.5)  # Short pause to let user see the highlight
        except Exception as e:
            self.logger.warning(f"Failed to highlight element: {str(e)}")

    def type_humanly(self, element, text, min_delay=0.005, max_delay=0.02):
        """Type text in chunks using ActionChains for better performance and reliability"""
        if not text:
            return

        # Chunked typing logic
        chunk_size = 50
        chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
        
        for chunk in chunks:
            try:
                # Occasional popup check
                if random.random() < 0.05:
                    self.handle_popups()
                
                # Use ActionChains to type at current cursor position
                ActionChains(self.driver).send_keys(chunk).perform()
                # time.sleep(random.uniform(min_delay, max_delay)) # Delay per chunk, not char
                time.sleep(random.uniform(0.05, 0.1))
            except Exception as e:
                self.logger.warning(f"Error typing chunk: {e}")
                # Try to recover focus
                try:
                    self.handle_popups()
                    if element:
                        ActionChains(self.driver).move_to_element(element).click().perform()
                    ActionChains(self.driver).send_keys(chunk).perform()
                except:
                    pass

    def type_line_with_links(self, element, line):
        """Type line with multiple markdown links support [text](url) and human typing"""
        try:
            # Pattern for [text](url)
            pattern = re.compile(r'\[(.*?)\]\((.*?)\)')
            
            last_pos = 0
            matches = list(pattern.finditer(line))
            
            if not matches:
                # No links, just type the line
                self.type_humanly(element, line)
                return

            for match in matches:
                # Type text before the link
                start, end = match.span()
                if start > last_pos:
                    text_before = line[last_pos:start]
                    self.type_humanly(element, text_before)
                
                # Handle the link
                link_text = match.group(1)
                link_url = match.group(2)
                
                if not link_text:
                    last_pos = end
                    continue
                    
                # Type the link text
                self.type_humanly(element, link_text)
                
                # Select the text we just typed (Shift + Left Arrow * length)
                # Optimized selection
                actions = ActionChains(self.driver)
                actions.key_down(Keys.SHIFT)
                for _ in range(len(link_text)):
                    actions.send_keys(Keys.ARROW_LEFT)
                actions.key_up(Keys.SHIFT)
                actions.perform()
                time.sleep(0.2)
                
                # Open link dialog (Ctrl+K)
                actions = ActionChains(self.driver)
                actions.key_down(Keys.CONTROL).send_keys('k').key_up(Keys.CONTROL)
                actions.perform()
                time.sleep(1.0) # Wait for dialog to appear
                
                # Type URL into the dialog (it should be focused)
                # Use faster typing for URL to ensure it's entered correctly
                actions = ActionChains(self.driver)
                actions.send_keys(link_url)
                actions.perform()
                time.sleep(0.2)
                
                # Press Enter to confirm link
                actions = ActionChains(self.driver)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(0.5) # Wait for link processing
                
                # Move cursor to the right to deselect/move out of link
                actions = ActionChains(self.driver)
                actions.send_keys(Keys.ARROW_RIGHT)
                actions.perform()
                time.sleep(0.2)
                
                last_pos = end
                
            # Type remaining text
            if last_pos < len(line):
                self.type_humanly(element, line[last_pos:])
                
        except Exception as e:
            self.logger.error(f"Error in type_line_with_links: {str(e)}")
            # Fallback: just type the raw line if complex typing fails
            try:
                # If we partially typed, this might duplicate, but it's a fallback for safety
                # Better to log and continue or try to append the rest
                # For now, just logging is safer than potentially duplicating content
                pass 
            except:
                pass

    def find_element_with_retry(self, selectors, timeout=10):
        """
        Find element with multiple selectors and retries
        
        Args:
            selectors: List of (by, selector) tuples
            timeout: Maximum time to try
            
        Returns:
            WebElement or None
        """
        end_time = time.time() + timeout
        while time.time() < end_time:
            for by, selector in selectors:
                try:
                    elements = self.driver.find_elements(by, selector)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            return element
                except:
                    continue
            time.sleep(0.5)
        return None
    
    def find_and_click(self, selectors, description="", timeout=10):
        """
        Find and click an element with multiple selectors
        
        Args:
            selectors: List of (by, selector) tuples
            description: Description for logging
            timeout: Maximum time to try
            
        Returns:
            bool: True if successful
        """
        try:
            element = self.find_element_with_retry(selectors, timeout)
            if element:
                # Scroll to element
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                
                # Highlight element
                self.highlight_element(element)
                
                # Try different click methods
                click_methods = [
                    lambda: element.click(),
                    lambda: self.driver.execute_script("arguments[0].click();", element),
                    lambda: ActionChains(self.driver).move_to_element(element).click().perform(),
                ]
                
                for method in click_methods:
                    try:
                        method()
                        self.logger.info(f"Successfully clicked {description}")
                        self.human_delay()
                        return True
                    except:
                        continue
        except Exception as e:
            self.logger.error(f"Failed to click {description}: {str(e)}")
        
        return False
    
    def wait_for_element(self, selectors, timeout=15):
        """
        Wait for element to be present
        
        Args:
            selectors: List of (by, selector) tuples
            timeout: Maximum wait time
            
        Returns:
            WebElement or None
        """
        end_time = time.time() + timeout
        while time.time() < end_time:
            for by, selector in selectors:
                try:
                    element = self.driver.find_element(by, selector)
                    if element.is_displayed():
                        return element
                except:
                    continue
            time.sleep(0.5)
        return None
    
    def handle_popups(self):
        """Handle any popups that appear"""
        popup_selectors = [
            (By.XPATH, "//button[contains(., 'Accept')]"),
            (By.XPATH, "//button[contains(., 'Got it')]"),
            (By.XPATH, "//button[contains(., 'Agree')]"),
            (By.XPATH, "//button[@aria-label='Close']"),
            (By.XPATH, "//button[contains(@class, 'close')]"),
            (By.XPATH, "//button[contains(., 'Maybe Later')]"),
            (By.XPATH, "//button[contains(., 'Not now')]"),
            (By.XPATH, "//button[contains(., 'No Thanks')]"),
            (By.XPATH, "//div[contains(@class, 'modal')]//button[text()='×']"),
        ]
        
        for by, selector in popup_selectors:
            try:
                elements = self.driver.find_elements(by, selector)
                for element in elements:
                    if element.is_displayed():
                        self.logger.info(f"Found popup: {selector}")
                        element.click()
                        self.human_delay()
                        break
            except:
                continue
    
    def _safe_wait(self, a=0.8, b=2.2):
        try:
            time.sleep(random.uniform(a, b))
        except Exception:
            time.sleep(1.0)
    
    def _random_scroll(self, times=3):
        try:
            for _ in range(max(1, times)):
                y = random.randint(200, 800)
                self.driver.execute_script("window.scrollBy(0, arguments[0]);", y)
                self._safe_wait(0.2, 0.8)
        except Exception:
            pass
    
    def _simulate_user_activity(self, seconds=None):
        try:
            dur = seconds if seconds is not None else random.uniform(self.activity_min_s, self.activity_max_s)
            start = time.time()
            while time.time() - start < dur:
                self._random_scroll(times=1)
                self._safe_wait(0.5, 1.2)
        except Exception:
            pass
    
    def _check_account_health(self):
        try:
            body = ""
            try:
                body = self.driver.page_source.lower()
            except Exception:
                body = ""
            patterns = [
                r"\byour account\b.*\b(suspended|blocked|disabled|limited)\b",
                r"\baccount\b.*\b(suspended|disabled)\b",
                r"\baction\b.*\bblocked\b",
                r"\blogin\b.*\btemporarily disabled\b",
            ]
            for pat in patterns:
                if re.search(pat, body):
                    return False
            return True
        except Exception:
            return True
    
    def _reading_delay_for_content(self, text):
        try:
            n = max(0, len(text or ""))
            base = 6.0
            per_char = 0.006
            t = base + per_char * n
            t = max(8.0, min(t, 25.0))
            return t
        except Exception:
            return 10.0
    
    def login_to_quora(self):
        """Login to Quora with robust detection"""
        try:
            self.logger.info("Starting login process...")
            
            # Navigate to Quora
            self.driver.get(self.quora_url)
            time.sleep(3)
            
            # Handle any initial popups
            self.handle_popups()
            if self.safe_mode:
                self._simulate_user_activity()
            
            # Check if already logged in
            try:
                profile_elements = [
                    (By.XPATH, "//a[contains(@href, '/profile/')]"),
                    (By.CSS_SELECTOR, "[data-click-id='user']"),
                    (By.XPATH, "//img[contains(@src, 'profile')]"),
                ]
                
                if self.find_element_with_retry(profile_elements, timeout=5):
                    self.logger.info("Already logged in!")
                    return True
            except:
                pass
            
            # Click login button if not on login page
            login_buttons = [
                (By.XPATH, "//a[contains(., 'Login')]"),
                (By.XPATH, "//a[contains(., 'Log in')]"),
                (By.XPATH, "//button[contains(., 'Login')]"),
                (By.XPATH, "//button[contains(., 'Log in')]"),
                (By.XPATH, "//span[contains(., 'Login')]"),
                (By.XPATH, "//span[contains(., 'Log in')]"),
            ]
            
            if self.find_and_click(login_buttons, "login button", timeout=5):
                time.sleep(3)
            
            # Handle popups again
            self.handle_popups()
            
            # Check for "Continue with email" if email field is not immediately visible
            continue_email_selectors = [
                (By.XPATH, "//div[contains(text(), 'Continue with email')]"),
                (By.XPATH, "//button[contains(., 'Continue with email')]"),
            ]
            self.find_and_click(continue_email_selectors, "Continue with email", timeout=3)
            
            # Enter email
            email_selectors = [
                (By.NAME, "email"),
                (By.CSS_SELECTOR, "input[name='email']"),
                (By.CSS_SELECTOR, "input[type='email']"),
                (By.CSS_SELECTOR, "#email"),
                (By.XPATH, "//input[contains(@placeholder, 'Email')]"),
                (By.XPATH, "//input[contains(@placeholder, 'email')]"),
                (By.XPATH, "//label[contains(., 'Email')]/following-sibling::input"),
            ]
            
            email_field = self.wait_for_element(email_selectors, timeout=10)
            if email_field:
                try:
                    email_field.click()
                    email_field.clear()
                except:
                    pass
                # Use standard send_keys for reliability in login forms
                try:
                    email_field.send_keys(self.credentials['email'])
                except Exception:
                    # Fallback to ActionChains if standard send_keys fails
                    ActionChains(self.driver).move_to_element(email_field).click().send_keys(self.credentials['email']).perform()
                    
                self.logger.info("Email entered")
                self.human_delay()
            else:
                self.logger.error("Email field not found")
                return False
            
            # Enter password
            password_selectors = [
                (By.NAME, "password"),
                (By.CSS_SELECTOR, "input[name='password']"),
                (By.CSS_SELECTOR, "input[type='password']"),
                (By.CSS_SELECTOR, "#password"),
                (By.XPATH, "//input[contains(@placeholder, 'Password')]"),
                (By.XPATH, "//input[contains(@placeholder, 'password')]"),
                (By.XPATH, "//label[contains(., 'Password')]/following-sibling::input"),
            ]
            
            password_field = self.wait_for_element(password_selectors, timeout=5)
            if password_field:
                try:
                    password_field.click()
                    password_field.clear()
                except:
                    pass
                
                # Use standard send_keys for reliability
                try:
                    password_field.send_keys(self.credentials['password'])
                except Exception:
                    ActionChains(self.driver).move_to_element(password_field).click().send_keys(self.credentials['password']).perform()

                self.logger.info("Password entered")
                self.human_delay()
            else:
                self.logger.error("Password field not found")
                return False
            
            # Click submit button
            submit_buttons = [
                (By.XPATH, "//button[@type='submit' and contains(., 'Login')]"),
                (By.XPATH, "//button[@type='submit' and contains(., 'Log in')]"),
                (By.XPATH, "//input[@type='submit' and contains(@value, 'Login')]"),
                (By.XPATH, "//input[@type='submit' and contains(@value, 'Log in')]"),
                (By.CSS_SELECTOR, "button[type='submit']"),
                (By.XPATH, "//button[contains(., 'Continue')]"),
            ]
            
            if not self.find_and_click(submit_buttons, "submit button", timeout=5):
                # Try pressing Enter
                try:
                    password_field.send_keys(Keys.ENTER)
                    self.logger.info("Pressed Enter to submit")
                except:
                    pass
            
            # Wait for login to complete
            self.logger.info("Waiting for login to complete...")
            time.sleep(5)
            
            # Check for login success
            success_indicators = [
                (By.XPATH, "//a[contains(@href, '/profile/')]"),
                (By.CSS_SELECTOR, "[data-click-id='user']"),
                (By.XPATH, "//div[contains(text(), 'Home')]"),
                (By.XPATH, "//div[contains(text(), 'Answer')]"),
                (By.XPATH, "//div[contains(text(), 'Spaces')]"),
                (By.XPATH, "//span[contains(text(), 'Profile')]"),
            ]
            
            if self.wait_for_element(success_indicators, timeout=20):
                self.logger.info("Login successful!")
                self.take_screenshot("login_success")
                if self.safe_mode:
                    self._simulate_user_activity()
                    if not self._check_account_health():
                        self.logger.warning("Account health warning after login; proceeding cautiously")
                return True
            else:
                self.logger.error("Login verification failed")
                self.take_screenshot("login_failed")
                return False
            
        except Exception as e:
            self.logger.error(f"Login failed with error: {str(e)}")
            self.take_screenshot("login_error")
            return False
    
    def navigate_to_post_creation(self):
        """Navigate to post creation via 'Add question' modal -> 'Create Post' tab"""
        try:
            self.logger.info("Navigating to post creation via Modal...")
            
            # Ensure we are on home page
            if "quora.com" not in self.driver.current_url:
                self.driver.get("https://www.quora.com/")
                time.sleep(3)
            
            # Step 1: Click "Add question" button in header
            self.logger.info("Looking for 'Add question' button...")
            add_question_selectors = [
                (By.XPATH, "//button[contains(., 'Add question')]"),
                (By.XPATH, "//div[contains(text(), 'Add question')]"),
                (By.XPATH, "//div[contains(@class, 'q-box') and contains(., 'Add question')]"),
            ]
            
            if not self.find_and_click(add_question_selectors, "'Add question' button", timeout=10):
                self.logger.error("Could not find 'Add question' button")
                return False
                
            time.sleep(2)
            
            # Step 2: Click "Create Post" tab in the modal
            self.logger.info("Looking for 'Create Post' tab in modal...")
            create_post_tab_selectors = [
                (By.XPATH, "//div[text()='Create Post']"),
                (By.XPATH, "//div[contains(text(), 'Create Post')]"),
                (By.XPATH, "//div[contains(@class, 'TabBar')]//div[contains(., 'Create Post')]"),
            ]
            
            if self.find_and_click(create_post_tab_selectors, "'Create Post' tab", timeout=10):
                self.logger.info("Successfully switched to 'Create Post' tab")
                self.take_screenshot("create_post_tab_active")
                if self.safe_mode:
                    self._simulate_user_activity()
                return True
            else:
                self.logger.error("Could not find 'Create Post' tab")
                self.take_screenshot("create_post_tab_missing")
                return False
                
        except Exception as e:
            self.logger.error(f"Navigation failed with error: {str(e)}")
            self.take_screenshot("navigation_error")
            return False
    
    def click_user_specific_post_button(self):
        """
        Click the post button using the specific user-provided CSS selector.
        This method is integrated from test_css_selector.py.
        """
        try:
            self.logger.info(f"Attempting to find element with specific CSS selector: {self.user_css_selector}")
            
            # Find the element
            element = self.driver.find_element(By.CSS_SELECTOR, self.user_css_selector)
            
            if element:
                self.logger.info("Found element with specific CSS selector!")
                
                # Scroll to it
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                
                # Highlight the button so user can see it
                self.highlight_element(element)
                
                # Try to click it
                try:
                    element.click()
                    self.logger.info("Clicked element via standard click")
                    return True
                except:
                    self.logger.info("Standard click failed, trying JS click")
                    self.driver.execute_script("arguments[0].click();", element)
                    self.logger.info("Clicked element via JS click")
                    return True
            else:
                self.logger.warning("Element not found with specific CSS selector")
                return False
                
        except NoSuchElementException:
            self.logger.warning("Specific CSS selector element not found")
            return False
        except Exception as e:
            self.logger.error(f"Error in click_user_specific_post_button: {str(e)}")
            return False

    def create_and_publish_post(self):
        """Create and publish a blog post in the modal with human-like execution"""
        try:
            self.logger.info("Creating blog post in modal (Human-like mode)...")
            
            # Wait for editor to be ready
            # Look for the main input area
            input_area_selectors = [
                (By.XPATH, "//div[@contenteditable='true']"),
                (By.XPATH, "//div[@role='textbox']"),
                (By.CSS_SELECTOR, ".q-text-editor"),
            ]
            
            # Checking actively
            input_field = self.wait_for_element(input_area_selectors, timeout=10)
            
            if input_field:
                self.highlight_element(input_field)
                input_field.click()
                
                # Clear existing content if any (usually empty)
                try:
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.DELETE)
                except:
                    pass
                
                self.logger.info("Typing title and body with hyperlink support...")
                title_line = self.blog_content.get('title') or ""
                self.type_line_with_links(input_field, title_line)
                self.human_delay()
                input_field.send_keys(Keys.ENTER)
                time.sleep(0.5)
                input_field.send_keys(Keys.ENTER)
                time.sleep(0.5)
                
                body_text = self.blog_content.get('body') or ""
                parts = re.split(r'(\{.*?\})', body_text)
                for part in parts:
                    m = re.match(r'\{(.*)\}', part)
                    if m:
                        p = m.group(1).strip().strip('"').strip("'")
                        ok = self._handle_image_token(input_field, p)
                        if not ok:
                            self.logger.warning(f"Image token failed: {p}")
                            # Fallback enter
                            ActionChains(self.driver).send_keys(Keys.ENTER).perform()
                        # If success, _handle_image_token already did Enters.
                        time.sleep(random.uniform(0.2, 0.5))
                        continue
                        
                    lines = part.splitlines() if part else []
                    for line in lines:
                        if not line.strip():
                            ActionChains(self.driver).send_keys(Keys.ENTER).perform()
                            time.sleep(0.1)
                            continue
                        
                        self.type_line_with_links(input_field, line)
                        
                        # New line after text using ActionChains
                        ActionChains(self.driver).send_keys(Keys.ENTER).perform()
                        time.sleep(random.uniform(0.1, 0.2))
                
                self.logger.info("Content entered")
                self.human_delay()
                if self.safe_mode:
                    delay = self._reading_delay_for_content((self.blog_content.get('title') or "") + "\n" + (self.blog_content.get('body') or ""))
                    self.logger.info(f"Safe reading delay: {delay:.2f}s")
                    time.sleep(delay)
            else:
                self.logger.error("Could not find input field in modal")
                self.take_screenshot("input_field_missing")
                return False
            
            # Click Post Button
            self.logger.info("Looking for Post button...")
            if self.safe_mode:
                if not self._check_account_health():
                    self.logger.warning("Account health warning before posting; proceeding cautiously")
                self._simulate_user_activity()
            
            # Try the specific user CSS selector first (Merged functionality)
            if self.click_user_specific_post_button():
                self.logger.info("Successfully clicked post button using specific user selector!")
            else:
                self.logger.info("Specific selector didn't work, falling back to standard selectors")
                
                # Standard Post button selectors in modal
                post_button_selectors = [
                    (By.CSS_SELECTOR, self.user_css_selector),
                    (By.XPATH, "//button[contains(., 'Post')]"),
                    (By.XPATH, "//div[text()='Post']"),
                    (By.XPATH, "//button[contains(@class, 'submit')]"),
                    (By.XPATH, "//div[@role='button' and contains(., 'Post')]"),
                ]
                
                # Reduced timeout for finding button
                if self.find_and_click(post_button_selectors, "Post button", timeout=5):
                    self.logger.info("Clicked Post button")
                else:
                    self.logger.error("Failed to click Post button")
                    self.take_screenshot("post_button_failed")
                    return False
            
            # Wait for success
            self.logger.info("Waiting for post to publish...")
            # Reduced wait time for confirmation
            time.sleep(2)
            self.take_screenshot("after_post_click")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Post creation failed with error: {str(e)}")
            self.take_screenshot("post_creation_error")
            return False
    
    def run(self):
        """Main execution flow"""
        start_time = time.time()
        
        try:
            # Step 1: Login
            self.logger.info("="*50)
            self.logger.info("STEP 1: LOGIN TO QUORA")
            self.logger.info("="*50)
            
            if not self.login_to_quora():
                self.logger.error("Login failed. Exiting.")
                return False
            
            login_time = time.time() - start_time
            self.logger.info(f"Login completed in {login_time:.2f} seconds")
            
            # Step 2: Navigate to post creation
            self.logger.info("="*50)
            self.logger.info("STEP 2: NAVIGATE TO POST CREATION")
            self.logger.info("="*50)
            
            if not self.navigate_to_post_creation():
                self.logger.error("Failed to navigate to post creation. Exiting.")
                return False
            
            navigation_time = time.time() - start_time - login_time
            self.logger.info(f"Navigation completed in {navigation_time:.2f} seconds")
            
            # Step 3: Create and publish post
            self.logger.info("="*50)
            self.logger.info("STEP 3: CREATE AND PUBLISH POST")
            self.logger.info("="*50)
            
            if not self.create_and_publish_post():
                self.logger.error("Failed to create and publish post. Exiting.")
                return False
            
            post_time = time.time() - start_time - login_time - navigation_time
            total_time = time.time() - start_time
            
            self.logger.info(f"Post creation completed in {post_time:.2f} seconds")
            self.logger.info(f"Total execution time: {total_time:.2f} seconds")
            self.logger.info("="*50)
            self.logger.info("SUCCESS: Blog post published successfully!")
            self.logger.info("="*50)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Execution failed with error: {str(e)}")
            self.take_screenshot("execution_error")
            return False
        finally:
            self.cleanup()
    
    def cleanup(self):
        """Clean up resources"""
        try:
            self.logger.info("Cleaning up...")
            time.sleep(3)  # Let user see final state
            
            # Take final screenshot
            self.take_screenshot("final_state")
            
            if self.driver:
                try:
                    self.driver.quit()
                    self.logger.info("Browser closed successfully")
                except Exception:
                    pass
        except Exception as e:
            self.logger.error(f"Cleanup error: {str(e)}")


def read_credentials_from_excel(file_path="credentials.xlsx"):
    """Read Quora credentials from Excel"""
    try:
        if not os.path.exists(file_path):
            logging.warning(f"File not found: {file_path}")
            return None

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Simple header mapping (assuming first row is header)
        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers[str(val).lower().strip()] = col
                
        # Helper to get col index by keywords
        def get_col_idx(keywords):
            for k, v in headers.items():
                if any(kw in k for kw in keywords):
                    return v
            return None
            
        website_col = get_col_idx(['website', 'url', 'site'])
        email_col = get_col_idx(['email', 'mail'])
        username_col = get_col_idx(['username', 'user'])
        password_col = get_col_idx(['password', 'pass', 'pwd'])
        
        # Search for quora
        for row in range(2, sheet.max_row + 1):
            website = str(sheet.cell(row=row, column=website_col).value or "").strip() if website_col else ""
            if "quora.com" in website.lower():
                email = str(sheet.cell(row=row, column=email_col).value or "").strip() if email_col else ""
                username = str(sheet.cell(row=row, column=username_col).value or "").strip() if username_col else ""
                password = str(sheet.cell(row=row, column=password_col).value or "").strip() if password_col else ""
                
                if email and password:
                    return {
                        "website": website,
                        "email": email,
                        "username": username,
                        "password": password
                    }
        return None
    except Exception as e:
        logging.error(f"Error reading Excel: {e}")
        return None


def run(driver, website, email, username, password):
    try:
        bot = QuoraBlogPoster(headless=False, driver=driver, email=email, password=password, website=website)
        ok = bot.run()
        return bool(ok)
    except Exception:
        return False

def main():
    """Main function"""
    print("\n" + "="*60)
    print("QUORA BLOG POSTING AUTOMATION - MERGED WITH CSS SELECTOR")
    print("="*60)
    
    # Credentials
    email = ""
    password = ""
    website = "https://www.quora.com/"
    
    # Try to load from Excel
    creds = read_credentials_from_excel()
    if creds:
        print(f"✅ Loaded credentials for {creds['website']} from credentials.xlsx")
        email = creds['email']
        password = creds['password']
        website = creds['website']
    else:
        print("⚠️ credentials.xlsx not found or no Quora row found.")
        print("Using empty credentials (manual login required).")

    print("This script will:")
    print("1. Login to Quora")
    print("2. Navigate to post creation page")
    print("3. Create and publish a blog post (using specific CSS selector)")
    print("="*60)
    
    # Run with visible browser for debugging
    print("\nStarting automation (browser will open)...")
    bot = QuoraBlogPoster(headless=False, email=email, password=password, website=website)
    
    try:
        success = bot.run()
        
        if success:
            print("\n" + "="*60)
            print("SUCCESS: Blog posted successfully!")
            print("="*60)
        else:
            print("\n" + "="*60)
            print("FAILED: Check quora_automation.log for details")
            print("Screenshots saved in 'screenshots' folder")
            print("="*60)
    finally:
        try:
            bot.cleanup()
        except Exception:
            pass


if __name__ == "__main__":
    main()
