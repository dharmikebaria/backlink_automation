import time
import logging
import platform
import os
import csv
import random
import traceback
import re
import urllib.parse
from enum import Enum
from dataclasses import dataclass, field
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from typing import Dict, List, Tuple, Any
import undetected_chromedriver as uc
from website_analyzer import WebsiteAnalyzer
from tumblr import TumblrHandler
from dev import DevToAutomation
from patreon import automate_patreon
from noon import HackernoonAutomation
from m2 import MediumSpecificHandler
from sub import SubstackHandler
from write import run as writerscafe_run
from quora import run as quora_run

import sys
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

def _uc_safe_del(self):
    try:
        self.quit()
    except Exception:
        pass
uc.Chrome.__del__ = _uc_safe_del
logger = logging.getLogger(__name__)

IS_WINDOWS = (platform.system().lower() == 'windows')
LOG_FILE = "automation.log"
RESULTS_FILE = "results.csv"
BLOG_FILE = "blog.txt"
BROWSER_OPTIONS = {
    "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "window_size": "1366,768",
    "headless": False,
    "chrome_version": 120
}
PAGE_LOAD_TIMEOUT = 30
ELEMENT_TIMEOUT = 10
EMAIL = ""
USERNAME = ""
PASSWORD = ""

def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILE, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

class ResultStatus(Enum):
    SUCCESS = "success"
    FAILED = "failed"
    PENDING = "pending"

@dataclass
class LoginResult:
    website: str
    status: ResultStatus
    timestamp: str
    message: str = ""
    details: Dict = field(default_factory=dict)
    screenshot_path: str = ""
    profile_reached: bool = False

def clean_text(text: str) -> str:
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text.strip())

def get_domain_from_url(url: str) -> str:
    try:
        parsed = urllib.parse.urlparse(url)
        domain = parsed.netloc
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain
    except:
        return ""

def is_login_page(driver) -> bool:
    try:
        page_url = driver.current_url.lower()
        page_source = driver.page_source.lower()
        page_title = driver.title.lower()
        login_indicators = [
            'login', 'signin', 'sign-in', 'log in', 'sign in',
            'password', 'email', 'username', 'authenticate',
            'account', 'auth', 'authentication'
        ]
        for indicator in login_indicators:
            if (indicator in page_url or indicator in page_source or indicator in page_title):
                return True
        login_elements = [
            "input[type='password']",
            "input[name*='password']",
            "input[type='email']",
            "input[name*='email']",
            "input[name*='username']"
        ]
        for selector in login_elements:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if elements and len(elements) > 0:
                    return True
            except:
                continue
        return False
    except:
        return False

def is_profile_page(driver) -> bool:
    try:
        page_url = driver.current_url.lower()
        page_source = driver.page_source.lower()
        page_title = driver.title.lower()
        profile_indicators = [
            'profile', 'dashboard', 'account', 'home', 'feed',
            'welcome', 'console', 'workspace', 'overview',
            'my account', 'user panel', 'settings', 'edit profile',
            'personal information'
        ]
        for indicator in profile_indicators:
            if (indicator in page_url or indicator in page_source or indicator in page_title):
                return True
        return False
    except:
        return False

def create_driver():
    """Create driver for Reddit (compatibility function)"""
    try:
        options = uc.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-notifications')
        options.add_argument('--disable-popup-blocking')
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        options.add_argument('--window-size=1366,768')
        
        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.notifications": 2
        }
        options.add_experimental_option("prefs", prefs)
        
        driver = uc.Chrome(options=options, headless=False)
        driver.set_page_load_timeout(30)
        
        # Add stealth JS
        stealth_js = """
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
        """
        driver.execute_script(stealth_js)
        
        return driver
    except Exception as e:
        logger.error(f"Driver creation error: {e}")
        raise

def navigate_to_reddit(driver, url="https://www.reddit.com/"):
    """Navigate to Reddit and wait for page to load"""
    try:
        logger.info("🌐 Navigating to Reddit...")
        driver.get(url)
        time.sleep(4)
        
        # Wait for page to be interactive
        wait = WebDriverWait(driver, 10)
        wait.until(lambda d: d.execute_script('return document.readyState') == 'complete')
        
        return True
    except Exception as e:
        logger.error(f"Navigation error: {e}")
        return False

def trigger_login_modal(driver):
    """Click on login button to trigger modal - SIMPLIFIED"""
    try:
        logger.info("🔍 Looking for login button...")
        
        # First, try to find and click the login button
        login_buttons = [
            "//button[.//span[text()='Log In']]",
            "//button[contains(., 'Log In')]",
            "//a[contains(., 'Log In')]",
            "//span[text()='Log In']/..",
            "//*[@data-testid='login-button']",
            "//button[@aria-label='Log In']",
            "//button[contains(@class, 'login')]"
        ]
        
        for button_xpath in login_buttons:
            try:
                elements = driver.find_elements(By.XPATH, button_xpath)
                for element in elements:
                    if element.is_displayed() and element.is_enabled():
                        logger.info(f"✅ Found login button: {button_xpath}")
                        
                        # Highlight for visibility
                        driver.execute_script("arguments[0].style.border='3px solid green';", element)
                        time.sleep(0.5)
                        
                        # Try to click
                        try:
                            element.click()
                        except:
                            driver.execute_script("arguments[0].click();", element)
                        
                        time.sleep(3)
                        driver.execute_script("arguments[0].style.border='';", element)
                        return True
            except:
                continue
        
        # If no button found, try direct login URL
        logger.info("🔄 No visible login button found, trying direct login URL...")
        driver.get("https://www.reddit.com/login")
        time.sleep(3)
        return True
        
    except Exception as e:
        logger.error(f"Error triggering login: {e}")
        return False

def fill_login_form(driver, username_or_email, password):
    """Fill Reddit login form - COMPLETELY REWRITTEN"""
    try:
        logger.info("🔐 Attempting to fill Reddit login form...")
        
        # Wait a bit for modal to appear
        time.sleep(3)
        
        # Take a screenshot for debugging
        try:
            driver.save_screenshot("reddit_login_current_view.png")
            logger.info("📸 Screenshot saved: reddit_login_current_view.png")
        except:
            pass
        
        # METHOD 1: DIRECT APPROACH BASED ON IMAGE DESCRIPTION
        logger.info("🔍 METHOD 1: Looking for specific fields from image...")
        
        # Based on your image, the fields have labels "Email or username *" and "Password *"
        # The asterisk is likely part of the label text
        
        # Try to find the username/email field by looking near the label
        username_field = None
        password_field = None
        
        # First, find all input fields on the page
        all_inputs = driver.find_elements(By.TAG_NAME, "input")
        logger.info(f"📋 Found {len(all_inputs)} total input fields on page")
        
        # Look for visible input fields
        visible_inputs = []
        for inp in all_inputs:
            try:
                if inp.is_displayed():
                    input_type = inp.get_attribute("type") or "text"
                    input_name = inp.get_attribute("name") or "no-name"
                    input_id = inp.get_attribute("id") or "no-id"
                    logger.info(f"  Input: type='{input_type}', name='{input_name}', id='{input_id}'")
                    visible_inputs.append(inp)
            except:
                continue
        
        logger.info(f"👀 {len(visible_inputs)} visible input fields")
        
        # STRATEGY 1: Look for specific Reddit login field IDs/names
        field_selectors = [
            # Username/Email field selectors
            {"type": "username", "selectors": [
                (By.ID, "loginUsername"),
                (By.NAME, "username"),
                (By.NAME, "user"),
                (By.NAME, "email"),
                (By.CSS_SELECTOR, "input[type='email']"),
                (By.CSS_SELECTOR, "input[autocomplete='username']"),
                (By.CSS_SELECTOR, "input[autocomplete='email']"),
                (By.XPATH, "//input[@placeholder='Email or username']"),
                (By.XPATH, "//input[@placeholder='Email']"),
                (By.XPATH, "//input[@placeholder='Username']")
            ]},
            # Password field selectors
            {"type": "password", "selectors": [
                (By.ID, "loginPassword"),
                (By.NAME, "password"),
                (By.NAME, "pass"),
                (By.CSS_SELECTOR, "input[type='password']"),
                (By.CSS_SELECTOR, "input[autocomplete='current-password']"),
                (By.CSS_SELECTOR, "input[autocomplete='password']"),
                (By.XPATH, "//input[@placeholder='Password']")
            ]}
        ]
        
        for field_type in field_selectors:
            for by, selector in field_type["selectors"]:
                try:
                    elements = driver.find_elements(by, selector)
                    for element in elements:
                        try:
                            if element.is_displayed() and element.is_enabled():
                                if field_type["type"] == "username" and not username_field:
                                    username_field = element
                                    logger.info(f"✅ Found {field_type['type']} field: {by}='{selector}'")
                                    break
                                elif field_type["type"] == "password" and not password_field:
                                    password_field = element
                                    logger.info(f"✅ Found {field_type['type']} field: {by}='{selector}'")
                                    break
                        except:
                            continue
                except:
                    continue
        
        # STRATEGY 2: Find fields by their position relative to labels
        if not username_field or not password_field:
            logger.info("🔍 STRATEGY 2: Looking for fields by label text...")
            
            # Find labels with specific text from your image
            label_texts = ["Email or username", "Password", "Email", "Username"]
            
            for label_text in label_texts:
                try:
                    # Find labels containing this text
                    labels = driver.find_elements(By.XPATH, f"//*[contains(text(), '{label_text}')]")
                    
                    for label in labels:
                        try:
                            if label.is_displayed():
                                logger.info(f"📝 Found label: '{label_text}'")
                                
                                # Try to find associated input field
                                # Method 1: Input with corresponding id
                                label_for = label.get_attribute("for")
                                if label_for:
                                    try:
                                        field = driver.find_element(By.ID, label_for)
                                        if field.is_displayed():
                                            if "password" in label_text.lower() and not password_field:
                                                password_field = field
                                                logger.info(f"✅ Found password field via label 'for' attribute")
                                            elif not username_field:
                                                username_field = field
                                                logger.info(f"✅ Found username field via label 'for' attribute")
                                    except:
                                        pass
                                
                                # Method 2: Input following the label
                                try:
                                    # Find the next input element after this label
                                    field = label.find_element(By.XPATH, "following::input[1]")
                                    if field.is_displayed():
                                        if "password" in label_text.lower() and not password_field:
                                            password_field = field
                                            logger.info(f"✅ Found password field following label")
                                        elif not username_field:
                                            username_field = field
                                            logger.info(f"✅ Found username field following label")
                                except:
                                    pass
                                    
                                # Method 3: Input inside the same parent/form
                                try:
                                    parent = label.find_element(By.XPATH, "..")
                                    fields_in_parent = parent.find_elements(By.TAG_NAME, "input")
                                    for field in fields_in_parent:
                                        if field.is_displayed():
                                            field_type = field.get_attribute("type") or ""
                                            if field_type == "password" and not password_field:
                                                password_field = field
                                                logger.info(f"✅ Found password field in same parent")
                                            elif field_type != "password" and not username_field:
                                                username_field = field
                                                logger.info(f"✅ Found username field in same parent")
                                except:
                                    pass
                        except:
                            continue
                except:
                    continue
        
        # STRATEGY 3: Find by form structure
        if not username_field or not password_field:
            logger.info("🔍 STRATEGY 3: Looking for form structure...")
            
            # Find all forms
            forms = driver.find_elements(By.TAG_NAME, "form")
            logger.info(f"📋 Found {len(forms)} forms")
            
            for form in forms:
                try:
                    if form.is_displayed():
                        # Get all inputs in this form
                        inputs = form.find_elements(By.TAG_NAME, "input")
                        text_inputs = []
                        password_inputs = []
                        
                        for inp in inputs:
                            try:
                                if inp.is_displayed():
                                    inp_type = inp.get_attribute("type") or "text"
                                    if inp_type == "password":
                                        password_inputs.append(inp)
                                    elif inp_type in ["text", "email", ""]:
                                        text_inputs.append(inp)
                            except:
                                continue
                        
                        if len(text_inputs) >= 1 and len(password_inputs) >= 1:
                            username_field = text_inputs[0]
                            password_field = password_inputs[0]
                            logger.info("✅ Found fields in form structure")
                            break
                except:
                    continue
        
        # STRATEGY 4: Last resort - find first two visible inputs
        if not username_field or not password_field:
            logger.info("🔍 STRATEGY 4: Looking for first two visible inputs...")
            
            visible_text_inputs = []
            visible_password_inputs = []
            
            for inp in visible_inputs:
                try:
                    inp_type = inp.get_attribute("type") or "text"
                    if inp_type == "password":
                        visible_password_inputs.append(inp)
                    else:
                        visible_text_inputs.append(inp)
                except:
                    continue
            
            if visible_text_inputs and visible_password_inputs:
                username_field = visible_text_inputs[0]
                password_field = visible_password_inputs[0]
                logger.info("✅ Using first visible text and password inputs")
        
        # Now fill the fields if found
        if username_field and password_field:
            logger.info("✍️ Filling login fields...")
            
            # Fill username/email
            try:
                username_field.clear()
                time.sleep(0.3)
                
                # Type slowly
                for char in username_or_email:
                    username_field.send_keys(char)
                    time.sleep(0.05)
                
                logger.info(f"📧 Entered username/email: {username_or_email[:10]}...")
            except Exception as e:
                logger.error(f"Error filling username: {e}")
                username_field.send_keys(username_or_email)
            
            time.sleep(1)
            
            # Fill password
            try:
                password_field.clear()
                time.sleep(0.3)
                
                # Type slowly
                for char in password:
                    password_field.send_keys(char)
                    time.sleep(0.05)
                
                logger.info("🔑 Entered password")
            except Exception as e:
                logger.error(f"Error filling password: {e}")
                password_field.send_keys(password)
            
            time.sleep(1)
            
            # Try to submit
            logger.info("🔘 Attempting to submit login form...")
            
            # Method 1: Press Enter on password field
            try:
                password_field.send_keys(Keys.ENTER)
                logger.info("✅ Pressed Enter on password field")
                time.sleep(3)
                return True
            except:
                pass
            
            # Method 2: Find and click submit button
            submit_buttons = [
                "//button[@type='submit']",
                "//button[contains(., 'Log In')]",
                "//button[contains(., 'Sign In')]",
                "//button[contains(., 'Login')]",
                "//button[contains(., 'Continue')]",
                "//input[@type='submit']",
                "//button[@data-testid='login']"
            ]
            
            for button_xpath in submit_buttons:
                try:
                    buttons = driver.find_elements(By.XPATH, button_xpath)
                    for button in buttons:
                        if button.is_displayed() and button.is_enabled():
                            logger.info(f"✅ Found submit button: {button_xpath}")
                            button.click()
                            time.sleep(3)
                            return True
                except:
                    continue
            
            # Method 3: JavaScript click on any button in the modal
            try:
                driver.execute_script("""
                    var buttons = document.querySelectorAll('button');
                    for (var i = 0; i < buttons.length; i++) {
                        var btn = buttons[i];
                        var text = btn.textContent || btn.innerText;
                        if (text && (text.includes('Log In') || text.includes('Sign In') || text.includes('Login') || text.includes('Continue'))) {
                            btn.click();
                            return true;
                        }
                    }
                    return false;
                """)
                logger.info("✅ Attempted JavaScript click on login button")
                time.sleep(3)
                return True
            except:
                pass
            
            # If we filled fields but couldn't submit, still return True
            # as the user might need to manually click
            logger.info("⚠️ Fields filled but could not auto-submit")
            return True
        
        else:
            logger.error(f"❌ Could not find required fields. Username field: {username_field is not None}, Password field: {password_field is not None}")
            
            # Take another screenshot to see what's on screen
            try:
                driver.save_screenshot("reddit_login_error.png")
                logger.info("📸 Error screenshot saved: reddit_login_error.png")
            except:
                pass
            
            return False
        
    except Exception as e:
        logger.error(f"❌ Error in fill_login_form: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def is_logged_in(driver):
    """Check if user is logged in to Reddit"""
    try:
        # Wait a bit for page to update
        time.sleep(5)
        
        # Check multiple indicators
        indicators = [
            # Logout button/link
            "//button[contains(., 'Log Out')]",
            "//a[contains(., 'Log Out')]",
            "//button[contains(., 'Logout')]",
            "//a[contains(., 'Logout')]",
            
            # User menu/avatar
            "//button[contains(@aria-label, 'Account')]",
            "//button[contains(@aria-label, 'User')]",
            "//div[contains(@class, 'UserMenu')]",
            "//button[contains(@class, 'UserMenu')]",
            "//img[contains(@alt, 'Avatar')]",
            
            # User-specific elements
            "//a[contains(@href, '/user/')]",
            "//span[contains(text(), 'u/')]",
            
            # Feed elements that only show when logged in
            "//div[contains(@class, 'home-container')]",
            "//div[contains(@class, 'feed')]"
        ]
        
        for indicator in indicators:
            try:
                elements = driver.find_elements(By.XPATH, indicator)
                for element in elements:
                    if element.is_displayed():
                        logger.info(f"✅ Logged in indicator found: {indicator}")
                        return True
            except:
                continue
        
        # Check page source for logout text
        page_source = driver.page_source.lower()
        logout_indicators = ["logout", "log out", "sign out"]
        for indicator in logout_indicators:
            if indicator in page_source:
                logger.info(f"✅ Found '{indicator}' in page source")
                return True
        
        # Check URL - if not on login page
        current_url = driver.current_url.lower()
        if "login" not in current_url and "signin" not in current_url:
            logger.info("✅ Not on login page - assuming logged in")
            return True
        
        logger.info("❌ No login indicators found")
        return False
        
    except Exception as e:
        logger.error(f"Error checking login status: {e}")
        return False

def read_blog_text():
    """Read blog text from file"""
    try:
        blog_file = "blog.txt"
        if os.path.exists(blog_file):
            with open(blog_file, 'r', encoding='utf-8') as f:
                content = f.read().strip()
            
            lines = content.split('\n')
            title = lines[0] if lines else f"Blog Post {time.strftime('%Y-%m-%d')}"
            body = '\n'.join(lines[1:]) if len(lines) > 1 else content
            
            return {
                'title': title[:300],
                'body': body[:40000]
            }
        else:
            return {
                'title': f"Automated Reddit Post {time.strftime('%Y-%m-%d %H:%M')}",
                'body': "This is an automated post created by the Universal Login Bot."
            }
    except Exception as e:
        logger.error(f"Error reading blog text: {e}")
        return {
            'title': "Automated Post",
            'body': "Content could not be loaded."
        }

def create_post(driver, username, title, body):
    """Create a post on Reddit"""
    try:
        logger.info("📝 Creating Reddit post...")
        
        # Navigate to submit page
        driver.get("https://www.reddit.com/submit")
        time.sleep(3)
        
        # Wait for page to load
        wait = WebDriverWait(driver, 10)
        
        # Check if we need to select a community
        try:
            community_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Choose a community']"))
            )
            community_input.click()
            time.sleep(1)
            
            # Select first available community
            first_community = driver.find_element(By.XPATH, "//div[contains(@class, 'community-list')]//div[@role='button']")
            first_community.click()
            time.sleep(1)
        except:
            logger.info("ℹ️ Community selection not found or not required")
        
        # Fill title
        title_field = wait.until(
            EC.presence_of_element_located((By.XPATH, "//textarea[@placeholder='Title']"))
        )
        title_field.clear()
        title_field.send_keys(title)
        time.sleep(1)
        
        # Fill body/text
        try:
            # Try to find text editor
            body_field = driver.find_element(By.XPATH, "//div[@role='textbox']")
            body_field.click()
            body_field.send_keys(body)
        except:
            # Try alternative selector
            try:
                body_field = driver.find_element(By.XPATH, "//textarea[@placeholder='Text (optional)']")
                body_field.send_keys(body)
            except:
                logger.warning("Could not find body field, posting with title only")
        
        time.sleep(2)
        
        # Submit post
        submit_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Post')]")
        submit_button.click()
        
        time.sleep(5)
        
        # Check if post was successful
        try:
            success_indicator = driver.find_element(By.XPATH, "//div[contains(text(), 'posted') or contains(text(), 'submitted')]")
            logger.info("✅ Post created successfully")
            return True
        except:
            # If no success indicator but URL changed, assume success
            if "comments" in driver.current_url:
                logger.info("✅ Post appears to be created (comments URL)")
                return True
            else:
                logger.warning("⚠️ Post submission status unclear")
                return False
        
    except Exception as e:
        logger.error(f"❌ Error creating post: {e}")
        return False

def install_packages():
    print("🔧 Installing required packages...")
    packages = [
        'undetected-chromedriver==3.5.4',
        'selenium==4.15.2',
        'requests==2.31.0',
        'openpyxl==3.1.2'
    ]
    for package in packages:
        try:
            package_name = package.split('==')[0]
            __import__(package_name.replace('-', '_'))
            print(f"✅ {package_name} already installed")
        except ImportError:
            print(f"📦 Installing {package}...")
            try:
                import subprocess, sys as _sys
                subprocess.check_call([_sys.executable, "-m", "pip", "install", package])
            except:
                print(f"❌ Failed to install {package}")
    print("\n✅ All packages installed!")

def cleanup_chromedriver():
    try:
        pass
    except:
        pass

def save_result_csv(result: LoginResult):
    try:
        file_exists = os.path.exists(RESULTS_FILE)
        with open(RESULTS_FILE, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['timestamp', 'website', 'status', 'message', 'profile_reached', 'final_url', 'screenshot'])
            writer.writerow([
                result.timestamp,
                result.website,
                result.status.value,
                result.message,
                result.profile_reached,
                result.details.get('final_url', '')[:100],
                result.screenshot_path
            ])
    except Exception as e:
        print(f"❌ Error saving result to CSV: {e}")

class UniversalLoginBot:
    def __init__(self):
        self.email = EMAIL
        self.username = USERNAME
        self.password = PASSWORD
        self.website_analyzer = WebsiteAnalyzer()
        self.driver = None
        self.wait = None
        self.tumblr_handler = None

    def _get_chrome_version(self):
        try:
            import subprocess
            result = subprocess.run(['reg', 'query', 'HKEY_CURRENT_USER\\Software\\Google\\Chrome\\BLBeacon', '/v', 'version'],
                                  capture_output=True, text=True, shell=True)
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    if 'version' in line.lower():
                        version = line.split()[-1]
                        major_version = int(version.split('.')[0])
                        return major_version
            result = subprocess.run(['chrome', '--version'], capture_output=True, text=True)
            if result.returncode == 0:
                version = result.stdout.strip().split()[-1]
                major_version = int(version.split('.')[0])
                return major_version
        except:
            pass
        return 120

    def _add_stealth_js(self):
        stealth_js = """
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
        """
        try:
            self.driver.execute_script(stealth_js)
        except:
            pass

    def create_driver(self):
        try:
            options = uc.ChromeOptions()
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--disable-notifications')
            options.add_argument('--disable-popup-blocking')
            options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            options.add_argument('--window-size=1366,768')
            prefs = {
                "credentials_enable_service": False,
                "profile.password_manager_enabled": False,
                "profile.default_content_setting_values.notifications": 2
            }
            options.add_experimental_option("prefs", prefs)
            driver = uc.Chrome(options=options, headless=False)
            self.driver = driver
            self.driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
            self.wait = WebDriverWait(self.driver, ELEMENT_TIMEOUT)
            self._add_stealth_js()
            self.tumblr_handler = TumblrHandler(self.driver)
            return True
        except Exception as e:
            logger.error(f"❌ Failed to create driver: {e}")
            if hasattr(self, 'driver') and self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass
            self.driver = None
            self.wait = None
            return False

    def navigate_to_login_page(self, url: str) -> bool:
        try:
            logger.info(f"🌐 Visiting: {url}")
            self.driver.get(url)
            time.sleep(5)
            if 'livejournal.com' in url.lower():
                live_ok = self._handle_livejournal_homepage()
                if live_ok:
                    return True
            if is_login_page(self.driver):
                return True
            login_button = self.website_analyzer.find_login_button(self.driver)
            if login_button:
                try:
                    login_button.click()
                    time.sleep(5)
                    if is_login_page(self.driver):
                        return True
                except Exception as e:
                    logger.error(f"❌ Failed to click login button: {e}")
            common_login_paths = [
                "/login",
                "/signin",
                "/auth/login",
                "/account/login",
                "/user/login",
                "/auth",
                "/sign-in"
            ]
            base_url = url.rstrip('/')
            for path in common_login_paths:
                try:
                    login_url = f"{base_url}{path}"
                    self.driver.get(login_url)
                    time.sleep(5)
                    if is_login_page(self.driver):
                        return True
                except:
                    continue
            return False
        except Exception as e:
            logger.error(f"❌ Navigation error: {e}")
            return False

    def _handle_livejournal_homepage(self) -> bool:
        try:
            self.driver.execute_script("window.scrollTo(0,0)")
            time.sleep(2)
            attempts = 2
            while attempts > 0:
                login_selectors = [
                    "//a[text()='LOG IN']",
                    "//button[text()='LOG IN']",
                    "//a[text()='Log In']",
                    "//button[text()='Log In']",
                    "//a[text()='Login']",
                    "//button[text()='Login']",
                    "//a[text()='Sign In']",
                    "//button[text()='Sign In']",
                    "//a[contains(text(),'SIGN IN')]",
                    "//button[contains(text(),'SIGN IN')]",
                    "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
                    "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
                    "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
                    "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
                    "//a[contains(text(),'LOGIN')]",
                    "//button[contains(text(),'LOGIN')]",
                    "header a[href*='login']",
                    "a[href*='/login']",
                    "a[href*='login.bml']"
                ]
                for selector in login_selectors:
                    try:
                        elements = self.driver.find_elements(By.XPATH, selector) if selector.startswith("//") else self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for element in elements:
                            if element.is_displayed() and element.is_enabled():
                                try:
                                    visible_text = (element.text or "").strip()
                                    inner_text = (self.driver.execute_script("return arguments[0].innerText;", element) or "").strip()
                                    text_norm = (visible_text or inner_text).lower()
                                    if ("log in" in text_norm) or ("login" in text_norm) or ("sign in" in text_norm) or ("signin" in text_norm):
                                        try:
                                            element.click()
                                        except:
                                            try:
                                                self.driver.execute_script("arguments[0].click();", element)
                                            except:
                                                try:
                                                    ActionChains(self.driver).move_to_element(element).click().perform()
                                                except:
                                                    pass
                                        time.sleep(5)
                                        if is_login_page(self.driver):
                                            return True
                                except Exception:
                                    continue
                    except:
                        continue
                self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                time.sleep(2)
                attempts -= 1
            return False
        except Exception as e:
            logger.error(f"❌ LiveJournal homepage handling error: {e}")
            return False

    def _submit_login_form(self, analysis: Dict) -> bool:
        try:
            for button_info in analysis["submit_buttons"]:
                try:
                    button = button_info["element"]
                    if button.is_displayed() and button.is_enabled():
                        button.click()
                        time.sleep(5)
                        return True
                except:
                    continue
            try:
                password_fields = [field["element"] for field in analysis["password_fields"]]
                if password_fields:
                    password_fields[0].send_keys(Keys.ENTER)
                    time.sleep(5)
                    return True
            except:
                pass
            submit_selectors = [
                "button[type='submit']",
                "input[type='submit']",
                "button:contains('Log In')",
                "button:contains('Sign In')",
                "button:contains('Login')",
                "button:contains('Continue')",
                "button:contains('Submit')"
            ]
            for selector in submit_selectors:
                try:
                    if 'contains' in selector:
                        text = selector.split("contains('")[1].split("')")[0]
                        xpath = f"//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]"
                        elements = self.driver.find_elements(By.XPATH, xpath)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            element.click()
                            time.sleep(5)
                            return True
                except:
                    continue
            return False
        except Exception as e:
            logger.error(f"❌ Form submission error: {e}")
            return False

    def _try_fallback_login(self) -> bool:
        try:
            all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
            visible_inputs = [inp for inp in all_inputs if inp.is_displayed() and inp.is_enabled()]
            if len(visible_inputs) >= 2:
                field1 = visible_inputs[0]
                field2 = visible_inputs[1]
                try:
                    field1.clear()
                    field1.send_keys(self.email)
                    time.sleep(1)
                    field2.clear()
                    field2.send_keys(self.password)
                    time.sleep(1)
                    field2.send_keys(Keys.ENTER)
                    time.sleep(5)
                    return True
                except:
                    pass
            login_selectors = [
                "button:contains('Log In')",
                "button:contains('Sign In')",
                "a:contains('Log In')",
                "a:contains('Sign In')"
            ]
            for selector in login_selectors:
                try:
                    text = selector.split("contains('")[1].split("')")[0]
                    xpath = f"//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]"
                    elements = self.driver.find_elements(By.XPATH, xpath)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            element.click()
                            time.sleep(5)
                            return True
                except:
                    continue
            return False
        except Exception as e:
            logger.error(f"Fallback login error: {e}")
            return False

    def fill_login_form(self, analysis: Dict) -> bool:
        try:
            current_url = self.driver.current_url.lower()
            if 'livejournal.com' in current_url:
                try:
                    username_candidates = []
                    password_candidates = []
                    selectors_user = [
                        "input[name='user']",
                        "input[name*='user']",
                        "input[name*='login']",
                        "input[id*='user']",
                        "input[id*='login']",
                        "input[type='text']",
                        "input[type='email']"
                    ]
                    selectors_pass = [
                        "input[name='password']",
                        "input[name*='pass']",
                        "input[id*='pass']",
                        "input[type='password']"
                    ]
                    for sel in selectors_user:
                        try:
                            for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                                if el.is_displayed() and el.is_enabled():
                                    t = (el.get_attribute("type") or "").lower()
                                    if t != "password":
                                        username_candidates.append(el)
                        except:
                            continue
                    for sel in selectors_pass:
                        try:
                            for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                                if el.is_displayed() and el.is_enabled():
                                    password_candidates.append(el)
                        except:
                            continue
                    u = username_candidates[0] if username_candidates else None
                    p = password_candidates[0] if password_candidates else None
                    if u and p:
                        try:
                            u.clear()
                        except:
                            pass
                        val_u = self.username or self.email
                        for ch in val_u:
                            u.send_keys(ch)
                            time.sleep(0.02)
                        time.sleep(0.3)
                        try:
                            p.clear()
                        except:
                            pass
                        for ch in self.password:
                            p.send_keys(ch)
                            time.sleep(0.02)
                        time.sleep(0.3)
                        try:
                            p.send_keys(Keys.ENTER)
                            time.sleep(4)
                            return True
                        except:
                            pass
                        btns = [
                            "//button[@type='submit']",
                            "//input[@type='submit']",
                            "//button[contains(., 'Log In')]",
                            "//button[contains(., 'Sign In')]",
                            "//button[contains(., 'Login')]",
                            "//button[contains(., 'Continue')]"
                        ]
                        for x in btns:
                            try:
                                for b in self.driver.find_elements(By.XPATH, x):
                                    if b.is_displayed() and b.is_enabled():
                                        b.click()
                                        time.sleep(4)
                                        return True
                            except:
                                continue
                    return False
                except:
                    return False
            if 'reddit.com' in current_url:
                return fill_login_form(self.driver, self.username or self.email, self.password)
            def get_element_from_field(field):
                if "element" in field:
                    return field["element"]
                try:
                    if field.get("id"):
                        return self.driver.find_element(By.ID, field["id"])
                    elif field.get("name"):
                        return self.driver.find_element(By.NAME, field["name"])
                    elif field.get("type") and field.get("placeholder"):
                        xpath = f"//input[@type='{field['type']}' and @placeholder='{field['placeholder']}']"
                        return self.driver.find_element(By.XPATH, xpath)
                except:
                    pass
                return None
            email_field = None
            if analysis["email_fields"]:
                for field in analysis["email_fields"]:
                    element = get_element_from_field(field)
                    if element and element.is_displayed() and element.is_enabled():
                        email_field = element
                        break
            username_field = None
            if not email_field and analysis["username_fields"]:
                for field in analysis["username_fields"]:
                    element = get_element_from_field(field)
                    if element and element.is_displayed() and element.is_enabled():
                        username_field = element
                        break
            password_field = None
            for field in analysis["password_fields"]:
                element = get_element_from_field(field)
                if element and element.is_displayed() and element.is_enabled():
                    password_field = element
                    break
            if (email_field or username_field) and password_field:
                if email_field:
                    try:
                        email_field.clear()
                        time.sleep(0.3)
                        for char in self.email:
                            email_field.send_keys(char)
                            time.sleep(0.05)
                    except:
                        email_field.send_keys(self.email)
                elif username_field:
                    try:
                        username_field.clear()
                        time.sleep(0.3)
                        for char in self.username:
                            username_field.send_keys(char)
                            time.sleep(0.05)
                    except:
                        username_field.send_keys(self.username)
                time.sleep(1)
                try:
                    password_field.clear()
                    time.sleep(0.3)
                    for char in self.password:
                        password_field.send_keys(char)
                        time.sleep(0.05)
                except:
                    password_field.send_keys(self.password)
                time.sleep(1)
                return self._submit_login_form(analysis)
            return False
        except Exception as e:
            logger.error(f"❌ Form filling error: {e}")
            return False

    def check_login_success(self) -> Tuple[bool, bool]:
        try:
            current_url = self.driver.current_url.lower()
            if 'reddit.com' in current_url:
                if is_logged_in(self.driver):
                    return True, True
            if 'livejournal.com' in current_url:
                try:
                    from live import is_logged_in as lj_is_logged_in
                    if lj_is_logged_in(self.driver):
                        return True, False
                    if is_login_page(self.driver):
                        return False, False
                except Exception:
                    pass
            error_selectors = [
                ".error", ".alert", ".message-error",
                "[class*='error']", "[class*='Error']",
                "div:contains('error')", "div:contains('Error')",
                "div:contains('invalid')", "div:contains('Invalid')",
                "div:contains('incorrect')", "div:contains('Incorrect')"
            ]
            for selector in error_selectors:
                try:
                    if 'contains' in selector:
                        text = selector.split("contains('")[1].split("')")[0]
                        xpath = f"//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]"
                        elements = self.driver.find_elements(By.XPATH, xpath)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        if element.is_displayed():
                            error_text = element.text.lower()
                            if any(keyword in error_text for keyword in ['error', 'invalid', 'incorrect', 'failed']):
                                return False, False
                except:
                    continue
            if is_login_page(self.driver):
                return False, False
            if is_profile_page(self.driver):
                return True, True
            success_indicators = [
                'dashboard', 'home', 'feed', 'profile', 'account',
                'welcome', 'console', 'workspace', 'overview',
                'main', 'news', 'timeline', 'stream'
            ]
            for indicator in success_indicators:
                if indicator in current_url:
                    return True, True
            page_title = self.driver.title.lower()
            for indicator in success_indicators:
                if indicator in page_title:
                    return True, True
            return True, False
        except Exception as e:
            logger.error(f"❌ Error checking login status: {e}")
            return False, False

    def navigate_to_profile(self) -> bool:
        try:
            profile_selectors = [
                "a[href*='profile']",
                "a[href*='account']",
                "a[href*='dashboard']",
                "a[href*='settings']",
                "a:contains('Profile')",
                "a:contains('Account')",
                "a:contains('Dashboard')",
                "a:contains('Settings')",
                "a:contains('My Account')",
                "button:contains('Profile')",
                "button:contains('Account')"
            ]
            for selector in profile_selectors:
                try:
                    if 'contains' in selector:
                        text = selector.split("contains('")[1].split("')")[0]
                        xpath = f"//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]"
                        elements = self.driver.find_elements(By.XPATH, xpath)
                    else:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            element.click()
                            time.sleep(5)
                            if is_profile_page(self.driver):
                                return True
                except:
                    continue
            current_url = self.driver.current_url
            base_url = current_url.split('/')[0] + '//' + self.driver.current_url.split('/')[2]
            profile_paths = [
                "/profile",
                "/account",
                "/dashboard",
                "/settings",
                "/user/profile",
                "/myaccount",
                "/home",
                "/feed"
            ]
            for path in profile_paths:
                try:
                    profile_url = f"{base_url}{path}"
                    self.driver.get(profile_url)
                    time.sleep(5)
                    if is_profile_page(self.driver):
                        return True
                except:
                    continue
            return False
        except Exception as e:
            logger.error(f"❌ Error navigating to profile: {e}")
            return False

    def handle_tumblr_post_login(self) -> bool:
        try:
            self.driver.get("https://www.tumblr.com/new/text")
            time.sleep(5)
        except Exception:
            pass
        try:
            return self.tumblr_handler.handle_post_login_actions()
        except Exception:
            return False

    def handle_reddit_post_login(self) -> bool:
        try:
            username = self.username or (self.email.split('@')[0] if self.email else '')
            blog_data = read_blog_text()
            title = blog_data.get('title', '')
            body = blog_data.get('body', '')
            return create_post(self.driver, username, title, body)
        except Exception:
            return False
    
    def handle_livejournal_post_login(self) -> bool:
        try:
            from live import create_post as lj_create_post, read_blog_text as lj_read_blog_text
            blog = lj_read_blog_text()
            title = blog.get('title', '') or "Automated LiveJournal Post"
            content = blog.get('body', '') or ""
            return lj_create_post(self.driver, title, content)
        except Exception:
            return False

    def login_to_website(self, url: str) -> LoginResult:
        logger.info(f"\n{'='*60}")
        logger.info(f"🚀 Attempting login to: {url}")
        logger.info(f"{'='*60}")
        from datetime import datetime
        result = LoginResult(
            website=url,
            status=ResultStatus.PENDING,
            timestamp=datetime.now().isoformat()
        )
        try:
            if not self.create_driver():
                result.status = ResultStatus.FAILED
                result.message = "Failed to create WebDriver"
                return result
            is_reddit = 'reddit.com' in url.lower()
            is_livejournal = 'livejournal.com' in url.lower()
            if not is_reddit:
                if not self.navigate_to_login_page(url):
                    result.status = ResultStatus.FAILED
                    result.message = "Could not find login page"
                    return result
            before_screenshot = self._take_screenshot("before_login")
            is_tumblr = 'tumblr' in url.lower()
            login_success = False
            profile_reached = False
            reddit_action_success = False
            if is_reddit:
                try:
                    self.driver.get(url)
                    time.sleep(4)
                    trigger_login_modal(self.driver)
                except Exception:
                    pass
                if fill_login_form(self.driver, self.username or self.email, self.password):
                    time.sleep(5)
                    if is_logged_in(self.driver):
                        login_success = True
                        profile_reached = True
                        reddit_action_success = self.handle_reddit_post_login()
                if not login_success:
                    result.status = ResultStatus.FAILED
                    result.message = "Reddit login failed"
                    return result
            else:
                analysis = self.website_analyzer.analyze_login_page(url, self.driver)
                if not analysis["email_fields"] and not analysis["username_fields"]:
                    if not self._try_fallback_login():
                        result.status = ResultStatus.FAILED
                        result.message = "No login form found"
                        return result
                else:
                    if not self.fill_login_form(analysis):
                        result.status = ResultStatus.FAILED
                        result.message = "Failed to fill/submit login form"
                        return result
                login_success, already_on_profile = self.check_login_success()
                profile_reached = already_on_profile
                if is_livejournal and login_success:
                    try:
                        lj_ok = self.handle_livejournal_post_login()
                    except Exception:
                        lj_ok = False
                else:
                    lj_ok = False
            tumblr_action_success = False
            if is_tumblr:
                tumblr_action_success = self.handle_tumblr_post_login()
                if tumblr_action_success:
                    login_success = True
            if not login_success and not (is_tumblr and tumblr_action_success):
                result.status = ResultStatus.FAILED
                result.message = "Login failed"
                return result
            if not profile_reached and not is_tumblr and not is_reddit:
                profile_reached = self.navigate_to_profile()
            after_screenshot = self._take_screenshot("after_login")
            result.status = ResultStatus.SUCCESS
            result.message = "Login successful"
            result.profile_reached = profile_reached
            result.screenshot_path = after_screenshot
            result.details = {
                "login_method": "reddit_handler" if is_reddit else "form",
                "profile_reached": profile_reached,
                "final_url": self.driver.current_url[:200],
                "final_title": self.driver.title[:100],
                "is_tumblr": is_tumblr,
                "tumblr_action_success": tumblr_action_success if is_tumblr else None,
                "is_reddit": is_reddit,
                "reddit_action_success": reddit_action_success if is_reddit else None,
                "is_livejournal": is_livejournal,
                "livejournal_action_success": lj_ok if is_livejournal else None
            }
            return result
        except Exception as e:
            logger.error(f"🔥 Error: {str(e)[:200]}")
            result.status = ResultStatus.FAILED
            result.message = f"Error: {str(e)[:200]}"
            try:
                if self.driver:
                    result.screenshot_path = self._take_screenshot("error")
            except:
                pass
            return result
        finally:
            self._close_driver()

    def _take_screenshot(self, prefix: str) -> str:
        try:
            os.makedirs("screenshots", exist_ok=True)
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            import string
            random_str = ''.join(random.choices(string.ascii_lowercase + string.digits, k=4))
            domain = "website"
            try:
                if self.driver:
                    parsed = urllib.parse.urlparse(self.driver.current_url)
                    domain = parsed.netloc.replace('.', '_')[:20]
            except:
                pass
            filename = f"{prefix}_{domain}_{timestamp}_{random_str}.png"
            path = f"screenshots/{filename}"
            self.driver.save_screenshot(path)
            return filename
        except Exception as e:
            logger.error(f"Screenshot failed: {e}")
            return ""

    def _close_driver(self):
        if self.driver:
            try:
                if hasattr(self.driver, 'service') and self.driver.service:
                    self.driver.service.stop()
                self.driver.quit()
            except Exception:
                pass
            finally:
                self.driver = None
                self.wait = None

def load_all_credentials_from_excel(path: str = 'credentials.xlsx') -> List[Dict[str, str]]:
    try:
        import openpyxl, re
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header_map = {}
        for col in range(1, ws.max_column+1):
            v = str(ws.cell(row=1, column=col).value or '').strip().lower()
            if not v:
                continue
            if v in ('website','url','site'):
                header_map['website'] = col
            elif v in ('email','e-mail','mail','email_selector'):
                header_map['email'] = col
            elif v in ('username','user','login','username_selector'):
                header_map['username'] = col
            elif v in ('password','pass','pwd','password_selector'):
                header_map['password'] = col
        if not (header_map.get('password') and header_map.get('website')):
            for row in range(2, ws.max_row+1):
                for col in range(1, ws.max_column+1):
                    val_raw = ws.cell(row=row, column=col).value
                    s = str(val_raw or '').strip()
                    if not s:
                        continue
                    sl = s.lower()
                    if not header_map.get('website') and (sl.startswith('http') or 'tumblr.com' in sl or 'livejournal.com' in sl or 'reddit.com' in sl or 'hackernoon.com' in sl):
                        header_map['website'] = col
                    if not header_map.get('email') and '@' in s and '.' in s:
                        header_map['email'] = col
                    # Prefer columns whose header suggests password
                    header_text = str(ws.cell(row=1, column=col).value or '').strip().lower()
                    if not header_map.get('password'):
                        if any(k in header_text for k in ('pass','password','pwd','password_selector')):
                            header_map['password'] = col
                        elif len(s) >= 6 and (' ' not in s) and not sl.startswith('http') and '@' not in s:
                            if not re.fullmatch(r'[A-Za-z0-9_.-]{3,32}', s):  # avoid typical usernames
                                header_map['password'] = col
                    if not header_map.get('username'):
                        if re.fullmatch(r'[A-Za-z0-9_.-]{3,32}', s):
                            header_map['username'] = col
        def val(row, key):
            c = header_map.get(key)
            return str(ws.cell(row=row, column=c).value or '').strip() if c else ''
        rows = []
        for row in range(2, ws.max_row+1):
            website = val(row,'website')
            email = val(row,'email')
            username = val(row,'username')
            password = val(row,'password')
            if not website:
                continue
            # If password missing or equals username/email, try to infer from columns labelled as password
            if (not password) or (password == username) or (password == email):
                # scan columns again preferring header that contains 'pass'
                for col in range(1, ws.max_column+1):
                    header_text = str(ws.cell(row=1, column=col).value or '').strip().lower()
                    cell_value = str(ws.cell(row=row, column=col).value or '').strip()
                    if not cell_value:
                        continue
                    if any(k in header_text for k in ('pass','password','pwd','password_selector')) and cell_value not in (username, email):
                        password = cell_value
                        break
            if not (password and (email or username)):
                continue
            if not username and email:
                prefix = (email.split('@')[0]).strip()
                if re.fullmatch(r'[A-Za-z0-9_.-]{3,32}', prefix):
                    username = prefix
            rows.append({
                'website': website,
                'email': email,
                'username': username,
                'password': password
            })
        return rows
    except Exception:
        return []

def batch_main(creds_list=None):
    print(f"DEBUG: batch_main called with {type(creds_list)}: {creds_list}")
    if platform.system().lower() == 'windows':
        try:
            import sys
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass
    print("=" * 60)
    print("UNIVERSAL WEBSITE LOGIN BOT")
    print("=" * 60)
    print(f"System: {platform.system()} {platform.release()}")
    print("=" * 60)
    try:
        install_packages()
    except Exception:
        pass
    os.makedirs("screenshots", exist_ok=True)
    
    if not creds_list:
        try:
            excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'credentials.xlsx')
            creds_list = load_all_credentials_from_excel(excel_path)
            if not creds_list:
                print("❌ No credentials found in Excel")
                return
            print(f"✅ Loaded {len(creds_list)} credentials from Excel")
        except Exception as e:
            print(f"❌ Failed to load credentials from Excel: {e}")
            return

    print(f"✅ Loaded {len(creds_list)} credentials for processing")
    results = []
    for i, row in enumerate(creds_list, 1):
        website = row.get('website')
        email = row.get('email')
        username = row.get('username')
        password = row.get('password')
        if website and not website.startswith('http'):
            website = 'https://' + website
        print(f"\n{'='*60}")
        print(f"🌐 Processing {i}/{len(creds_list)}: {website}")
        print(f"{'='*60}")
        print(f"🔐 Using creds: email='{email}', username='{username}', password_len={len(password)}", flush=True)
        is_reddit = 'reddit' in (website or '').lower()
        is_dev = 'dev.to' in (website or '').lower()
        is_patreon = 'patreon' in (website or '').lower()
        
        if is_dev:
            print("➡️ Routing to Dev.to automation", flush=True)
            try:
                automation = DevToAutomation(email, password)
                automation.run_automation()
                result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Dev.to post successful")
                results.append(result)
                save_result_csv(result)
            except Exception as e:
                print(f"❌ Dev.to error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Dev.to error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'livejournal' in (website or '').lower():
            print("➡️ Routing to LiveJournal automation", flush=True)
            try:
                from live import run as livejournal_run
            except Exception as e:
                print(f"❌ LiveJournal module load error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="LiveJournal module not available")
                results.append(result)
                save_result_csv(result)
                continue
            try:
                driver = create_driver()
            except Exception as e:
                print(f"❌ Failed to create driver: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Driver creation failed")
                results.append(result)
                save_result_csv(result)
                continue
            try:
                ok = livejournal_run(driver, website, email, username, password)
                if ok:
                    result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="LiveJournal post successful", profile_reached=True)
                else:
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="LiveJournal post failed")
                results.append(result)
                save_result_csv(result)
            except Exception as e:
                print(f"❌ LiveJournal error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"LiveJournal error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
            finally:
                try:
                    driver.quit()
                except Exception:
                    pass
        elif 'tumblr' in (website or '').lower():
            print("➡️ Routing to Tumblr automation", flush=True)
            try:
                driver = create_driver()
                try:
                    handler = TumblrHandler(driver)
                    if handler.execute_login(email, password):
                         if handler.handle_post_login_actions():
                             result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Tumblr post successful", profile_reached=True)
                         else:
                             result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Tumblr post failed")
                    else:
                        result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Tumblr login failed")
                    
                    results.append(result)
                    save_result_csv(result)
                except Exception as e:
                    print(f"❌ Tumblr automation error: {e}")
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Tumblr error: {str(e)[:200]}")
                    results.append(result)
                    save_result_csv(result)
                finally:
                    try:
                        driver.quit()
                    except Exception:
                        pass
            except Exception as e:
                 print(f"❌ Tumblr driver error: {e}")
                 result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Tumblr driver error: {str(e)[:200]}")
                 results.append(result)
                 save_result_csv(result)
        elif is_patreon:
            print("➡️ Routing to Patreon automation", flush=True)
            try:
                automate_patreon(email, password)
                result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Patreon automation finished")
                results.append(result)
                save_result_csv(result)
            except Exception as e:
                print(f"❌ Patreon error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Patreon error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'substack' in (website or '').lower():
            print("➡️ Routing to Substack automation", flush=True)
            try:
                from sub import run as substack_run
                driver = create_driver()
                ok = substack_run(driver, website, email, username, password)
                if ok:
                    result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Substack post successful")
                else:
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Substack post failed")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ Substack error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Substack error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'medium' in (website or '').lower():
            print("➡️ Routing to Medium automation", flush=True)
            try:
                driver = create_driver()
                try:
                    from m2 import run as medium_run
                except Exception as e:
                    print(f"❌ Medium module load error: {e}")
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Medium module not available")
                    results.append(result)
                    save_result_csv(result)
                    try:
                        driver.quit()
                    except:
                        pass
                    continue
                ok = medium_run(driver, website, email, username, password)
                if ok:
                    result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Medium attempt finished")
                else:
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Medium attempt failed")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ Medium error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Medium error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'hackernoon' in (website or '').lower():
            print("➡️ Routing to Hackernoon automation", flush=True)
            try:
                from noon import run as hackernoon_run
                driver = create_driver()
                ok = hackernoon_run(driver, website, email, username, password)
                if ok:
                    result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Hackernoon run finished")
                else:
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Hackernoon run failed")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ Hackernoon error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Hackernoon error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'quora' in (website or '').lower():
            print("➡️ Routing to Quora automation", flush=True)
            try:
                driver = create_driver()
                ok = quora_run(driver, website, email, username, password)
                if ok:
                    result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Quora run finished")
                else:
                    result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Quora run failed")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ Quora error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Quora error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif 'writerscafe' in (website or '').lower():
            print("➡️ Routing to WritersCafe automation", flush=True)
            try:
                driver = create_driver()
                writerscafe_run(driver, website, email, username, password)
                result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="WritersCafe run finished")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ WritersCafe error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"WritersCafe error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        elif is_reddit:
            print("➡️ Routing to Reddit automation", flush=True)
            try:
                driver = create_driver()
                navigate_to_reddit(driver)
                trigger_login_modal(driver)
                if fill_login_form(driver, email, password):
                     result = LoginResult(website=website, status=ResultStatus.SUCCESS, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Reddit login attempted")
                else:
                     result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Reddit login failed")
                results.append(result)
                save_result_csv(result)
                try:
                    driver.quit()
                except:
                    pass
            except Exception as e:
                print(f"❌ Reddit error: {e}")
                result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message=f"Reddit error: {str(e)[:200]}")
                results.append(result)
                save_result_csv(result)
        else:
            print(f"⚠️ Unknown website: {website}")
            result = LoginResult(website=website, status=ResultStatus.FAILED, timestamp=time.strftime("%Y-%m-%dT%H:%M:%S"), message="Unknown website")
            results.append(result)
            save_result_csv(result)
        if i < len(creds_list):
            wait_time = random.randint(5, 10)
            print(f"\n⏳ Waiting {wait_time} seconds...")
            time.sleep(wait_time)
    print("\n" + "="*60)
    print("📊 FINAL REPORT")
    print("="*60)
    total = len(results)
    success = sum(1 for r in results if getattr(r, 'status', None) and getattr(r.status, 'value', '') == 'success')
    success_with_profile = sum(1 for r in results if getattr(r, 'status', None) and getattr(r.status, 'value', '') == 'success' and getattr(r, 'profile_reached', False))
    print(f"Total Websites Tested: {total}")
    print(f"Successful Logins: {success} ({(success/total*100 if total else 0):.1f}%)")
    print(f"Profile Page Reached: {success_with_profile} ({(success_with_profile/total*100 if total else 0):.1f}%)")
    print(f"Failed Logins: {total - success} ({(((total-success)/total*100) if total else 0):.1f}%)")
    print("\n✅ SUCCESSFUL LOGINS:")
    any_success = False
    for r in results:
        if getattr(r, 'status', None) and getattr(r.status, 'value', '') == 'success':
            any_success = True
            print(f"  ✓ {getattr(r, 'website', '')}")
    if not any_success:
        print("  None")
    print("\n❌ FAILED LOGINS:")
    any_failed = False
    for r in results:
        if getattr(r, 'status', None) and getattr(r.status, 'value', '') == 'failed':
            any_failed = True
            print(f"  ✗ {getattr(r, 'website', '')}")
    if not any_failed:
        print("  None")
    print("="*60)

def main():
    """Main function for standalone Reddit testing"""
    import sys
    
    print("🔴 REDDIT LOGIN TEST")
    print("=" * 50)
    
    # Get credentials
    email = input("Email/Username: ").strip()
    password = input("Password: ").strip()
    
    if not email or not password:
        print("❌ Credentials required")
        return
    
    # Create driver
    try:
        driver = create_driver()
    except Exception as e:
        print(f"❌ Failed to create driver: {e}")
        return
    
    try:
        # Navigate to Reddit
        print("🌐 Navigating to Reddit...")
        driver.get("https://www.reddit.com")
        time.sleep(4)
        
        # Click login button
        print("🔍 Clicking login button...")
        login_success = trigger_login_modal(driver)
        
        if not login_success:
            print("❌ Could not trigger login modal")
            driver.quit()
            return
        
        print("⏳ Waiting for login form...")
        time.sleep(3)
        
        # Fill login form
        print("🔐 Filling login form...")
        if fill_login_form(driver, email, password):
            print("✅ Login form filled")
            time.sleep(5)
            
            # Check login status
            if is_logged_in(driver):
                print("✅✅✅ LOGIN SUCCESSFUL!")
                
                # Ask if user wants to create a post
                create_post_choice = input("Create a post? (y/n): ").strip().lower()
                if create_post_choice == 'y':
                    blog_data = read_blog_text()
                    if create_post(driver, email, blog_data['title'], blog_data['body']):
                        print("✅ Post created successfully!")
                    else:
                        print("❌ Failed to create post")
            else:
                print("❌ Login failed - not logged in")
        else:
            print("❌ Failed to fill login form")
    
    except KeyboardInterrupt:
        print("\n⏹️ Stopped by user")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            driver.quit()
            print("✅ Driver closed")
        except:
            pass

if __name__ == "__main__":
    import sys
    import argparse
    import json
    
    # Fix encoding for Windows console/subprocess
    if sys.platform.startswith('win'):
        try:
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
        except Exception:
            pass

    parser = argparse.ArgumentParser(description='Automation Script')
    parser.add_argument('--creds-json', help='Path to JSON file containing credentials list')
    
    # Allow unknown args to pass (backward compatibility)
    args, unknown = parser.parse_known_args()
    
    if len(sys.argv) > 1 and sys.argv[1].lower() in ("test","reddit-test"):
        main()
    elif args.creds_json and os.path.exists(args.creds_json):
        try:
            with open(args.creds_json, 'r', encoding='utf-8') as f:
                creds = json.load(f)
            print(f"✅ Loaded {len(creds)} credentials from {args.creds_json}")
            batch_main(creds)
        except Exception as e:
            print(f"❌ Failed to load credentials from JSON: {e}")
    else:
        try:
            excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'credentials.xlsx')
            creds = load_all_credentials_from_excel(excel_path)
            print(f"✅ Loaded {len(creds)} credentials from Excel")
            batch_main(creds)
        except Exception as e:
            print("❌ Please provide --creds-json argument with path to credentials file, or ensure credentials.xlsx is available.")

