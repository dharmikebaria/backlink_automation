import os
import time
import logging
import re
import random
import subprocess
from typing import Optional, Dict
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

LOG_FILE = "automation.log"
BLOG_FILE = "blog.txt"

BROWSER_OPTIONS = {
    "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "window_size": "1366,768",
    "headless": False,
    "chrome_version": 120,
}

PAGE_LOAD_TIMEOUT = 30
ELEMENT_TIMEOUT = 15

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def install_packages():
    pkgs = [
        'undetected-chromedriver==3.5.4',
        'selenium==4.15.2',
        'openpyxl==3.1.2',
    ]
    for p in pkgs:
        try:
            __import__(p.split('==')[0].replace('-', '_'))
        except Exception:
            try:
                import sys, subprocess
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', p])
            except Exception:
                pass


def read_blog_text() -> Dict[str, str]:
    # यह सुनिश्चित करने के लिए कि blog.txt सही से पढ़ा जा रहा है
    try:
        if os.path.exists(BLOG_FILE):
            with open(BLOG_FILE, 'r', encoding='utf-8') as f:
                content_all = f.read().strip()
                if not content_all:
                    return {"title": "AI in Telecom", "body": "Default Content"}
                
                lines = content_all.splitlines()
                title = lines[0].strip()
                body = "\n".join(lines[1:]).strip()
                return {"title": title, "body": body}
    except Exception as e:
        logger.error(f"Error reading blog file: {e}")
    return {"title": "Automated Post", "body": "Content missing"}


def load_livejournal_credentials(path: str = 'credentials.xlsx') -> Optional[Dict[str, str]]:
    try:
        import openpyxl, re
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header_map = {}
        for col in range(1, ws.max_column + 1):
            v = str(ws.cell(row=1, column=col).value or '').strip().lower()
            if not v:
                continue
            if v in ('website', 'url', 'site'):
                header_map['website'] = col
            elif v in ('email', 'e-mail', 'mail'):
                header_map['email'] = col
            elif v in ('username', 'user', 'login'):
                header_map['username'] = col
            elif v in ('password', 'pass', 'pwd'):
                header_map['password'] = col
        def val(row, key):
            c = header_map.get(key)
            return str(ws.cell(row=row, column=c).value or '').strip() if c else ''
        for row in range(2, ws.max_row + 1):
            website = val(row, 'website')
            email = val(row, 'email')
            username = val(row, 'username')
            password = val(row, 'password')
            domain = (website or '').lower()
            if 'livejournal.com' in domain:
                if not username and email:
                    prefix = (email.split('@')[0]).strip()
                    if re.fullmatch(r'[A-Za-z0-9_.-]{3,32}', prefix):
                        username = prefix
                return {
                    'website': website,
                    'email': email,
                    'username': username,
                    'password': password,
                }
        return None
    except Exception:
        return None


def create_driver() -> Optional[uc.Chrome]:
    try:
        options = uc.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-notifications')
        options.add_argument('--disable-popup-blocking')
        options.add_argument(f"--user-agent={BROWSER_OPTIONS['user_agent']}")
        options.add_argument(f"--window-size={BROWSER_OPTIONS['window_size']}")
        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.notifications": 2,
        }
        options.add_experimental_option("prefs", prefs)
        driver = uc.Chrome(
            options=options,
            headless=BROWSER_OPTIONS['headless'],
            use_subprocess=False,
            version_main=BROWSER_OPTIONS['chrome_version'],
        )
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        WebDriverWait(driver, ELEMENT_TIMEOUT)
        return driver
    except Exception as e:
        logger.error(f"Driver error: {e}")
        return None


def screenshot(driver, prefix: str) -> str:
    try:
        os.makedirs('screenshots', exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fn = f"{prefix}_www_livejournal_com_{ts}.png"
        path = os.path.join('screenshots', fn)
        driver.save_screenshot(path)
        return path
    except Exception:
        return ''


def click_login_button(driver) -> bool:
    """LiveJournal के मुख्य पेज पर login button ढूंढे और click करे"""
    logger.info("🔍 LiveJournal के मुख्य पेज पर login button ढूंढ रहा हूं...")
    
    # LiveJournal specific selectors - image में दिख रहे button के लिए
    sels = [
        # सीधे "GET STARTED" button जो image में दिख रहा है
        "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'get started')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'get started')]",
        
        # "LOG IN" text वाले buttons/links
        "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
        
        # Sign in variations
        "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
        
        # Login text variations
        "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login')]",
        "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login')]",
        
        # Login via href attributes
        "a[href*='login.bml']",
        "a[href*='/login']",
        "a[href*='signin']",
        
        # CSS classes that might contain login buttons
        ".login-button",
        ".signin-button",
        "[class*='login']",
        "[class*='signin']",
        
        # Header navigation में login links
        "header a[href*='login']",
        "nav a[href*='login']",
    ]
    
    for sel in sels:
        try:
            if sel.startswith('//'):
                els = driver.find_elements(By.XPATH, sel)
            elif ':contains' in sel:
                # Extract text from contains selector
                if "contains('" in sel:
                    txt = sel.split("contains('")[1].split("')")[0]
                    xpath = f"//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'), '{txt.lower()}')]"
                    els = driver.find_elements(By.XPATH, xpath)
                else:
                    continue
            else:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
            
            for el in els:
                try:
                    if el.is_displayed() and el.is_enabled():
                        logger.info(f"✅ LiveJournal login button found: {sel}")
                        logger.info(f"   Button text: '{el.text}'")
                        
                        # Scroll into view
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(1)
                        
                        # Highlight for debugging
                        driver.execute_script("arguments[0].style.border='3px solid red';", el)
                        time.sleep(0.5)
                        
                        # Click with multiple methods
                        clicked = False
                        
                        # Method 1: Standard click
                        try:
                            el.click()
                            logger.info("✅ Login button clicked (standard)")
                            clicked = True
                        except Exception as e1:
                            logger.debug(f"Standard click failed: {e1}")
                        
                        # Method 2: JavaScript click
                        if not clicked:
                            try:
                                driver.execute_script("arguments[0].click();", el)
                                logger.info("✅ Login button clicked (JavaScript)")
                                clicked = True
                            except Exception as e2:
                                logger.debug(f"JS click failed: {e2}")
                        
                        # Method 3: ActionChains click
                        if not clicked:
                            try:
                                actions = ActionChains(driver)
                                actions.move_to_element(el).click().perform()
                                logger.info("✅ Login button clicked (ActionChains)")
                                clicked = True
                            except Exception as e3:
                                logger.debug(f"ActionChains failed: {e3}")
                        
                        # Remove highlight
                        driver.execute_script("arguments[0].style.border='';", el)
                        
                        if clicked:
                            time.sleep(3)  # Wait for login page to load
                            
                            # Verify we're now on login page
                            current_url = driver.current_url.lower()
                            page_source = driver.page_source.lower()
                            
                            login_indicators = ['login', 'signin', 'auth', 'password', 'username']
                            if any(indicator in current_url or indicator in page_source for indicator in login_indicators):
                                logger.info("✅ Successfully navigated to LiveJournal login page")
                                return True
                            else:
                                logger.info("⚠️ Clicked but not sure if on login page")
                except Exception as e:
                    logger.debug(f"Element error: {e}")
                    continue
        except Exception:
            continue
    
    return False


def fill_login_form(driver, username: str, password: str) -> bool:
    """Fill LiveJournal login form"""
    try:
        logger.info("🔍 LiveJournal login form ढूंढ रहा हूं...")
        
        # LiveJournal specific field selectors
        user_fields = [
            ("input[name='user']", "css"),
            ("input[id='username']", "css"),
            ("input[name='username']", "css"),
            ("input[placeholder*='Username']", "css"),
            ("input[placeholder*='username']", "css"),
            ("//input[@name='user']", "xpath"),
            ("//input[contains(@placeholder,'Username') or contains(@placeholder,'username')]", "xpath"),
        ]
        
        user_el = None
        for sel, kind in user_fields:
            try:
                els = driver.find_elements(By.XPATH, sel) if kind == 'xpath' else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed():
                        user_el = el
                        logger.info(f"✅ Username field found: {sel}")
                        break
                if user_el:
                    break
            except Exception as e:
                logger.debug(f"Selector {sel} failed: {e}")
                continue
        
        # For LiveJournal, sometimes it's email instead of username
        if not user_el:
            email_fields = [
                ("input[type='email']", "css"),
                ("input[name='email']", "css"),
                ("input[placeholder*='Email']", "css"),
            ]
            for sel, kind in email_fields:
                try:
                    els = driver.find_elements(By.XPATH, sel) if kind == 'xpath' else driver.find_elements(By.CSS_SELECTOR, sel)
                    for el in els:
                        if el.is_displayed():
                            user_el = el
                            logger.info(f"✅ Email field found (using as username): {sel}")
                            break
                    if user_el:
                        break
                except Exception:
                    continue
        
        password_fields = [
            ("input[type='password']", "css"),
            ("input[name='password']", "css"),
            ("input[id='password']", "css"),
            ("//input[@type='password']", "xpath"),
        ]
        
        pass_el = None
        for sel, kind in password_fields:
            try:
                els = driver.find_elements(By.XPATH, sel) if kind == 'xpath' else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed():
                        pass_el = el
                        logger.info(f"✅ Password field found: {sel}")
                        break
                if pass_el:
                    break
            except Exception:
                continue
        
        if not (user_el and pass_el):
            logger.error("❌ Login form fields not found")
            return False
        
        # Fill username/email
        try:
            user_el.click()
            time.sleep(0.3)
            user_el.clear()
            user_el.send_keys(username)
            logger.info(f"✅ Username/Email entered: {username}")
            time.sleep(0.5)
        except Exception as e:
            logger.error(f"Username field error: {e}")
            return False
        
        # Fill password
        try:
            pass_el.click()
            time.sleep(0.3)
            pass_el.clear()
            pass_el.send_keys(password)
            logger.info("✅ Password entered")
            time.sleep(0.5)
        except Exception as e:
            logger.error(f"Password field error: {e}")
            return False
        
        # Find and click submit button
        submit_buttons = [
            "input[type='submit']",
            "button[type='submit']",
            "//input[@value='Log In' or @value='Login']",
            "//button[contains(.,'Log In') or contains(.,'Login')]",
            "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
        ]
        
        for sel in submit_buttons:
            try:
                els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        logger.info(f"✅ Submit button found: {sel}")
                        
                        # Scroll and click
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.5)
                        
                        try:
                            el.click()
                            logger.info("✅ Submit button clicked")
                        except Exception:
                            driver.execute_script("arguments[0].click();", el)
                        
                        time.sleep(4)
                        return True
            except Exception:
                continue
        
        # If no submit button found, press Enter
        logger.info("🔄 No submit button found, pressing Enter...")
        pass_el.send_keys(Keys.ENTER)
        time.sleep(4)
        return True
        
    except Exception as e:
        logger.error(f"❌ LiveJournal login form error: {e}")
        return False


def is_logged_in(driver) -> bool:
    try:
        txt = driver.page_source.lower()
        if 'logout' in txt or 'log out' in txt:
            return True
        sels = ["a[href*='logout']", "//a[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log out')]"]
        for sel in sels:
            try:
                els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                if any(e.is_displayed() for e in els):
                    return True
            except Exception:
                continue
        
        # Check for user profile elements
        profile_selectors = [
            "//a[contains(@href,'profile')]",
            "//a[contains(@href,'account')]",
            "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'my journal')]",
            "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'dashboard')]",
        ]
        
        for sel in profile_selectors:
            try:
                els = driver.find_elements(By.XPATH, sel)
                if any(e.is_displayed() for e in els):
                    return True
            except Exception:
                continue
        
        return False
    except Exception:
        return False


def human_type(element, text, min_delay=0.01, max_delay=0.05):
    """Simulate human typing"""
    try:
        # Try sending all at once first as it is often more reliable for hidden inputs, 
        # but for contenteditable we might want char by char if the user requested "slow"
        # The user requested "typing ko bhi hlka slow karo", so let's do char-by-char for visible text
        for ch in text:
            try:
                element.send_keys(ch)
                time.sleep(random.uniform(min_delay, max_delay))
            except Exception:
                # If char by char fails, try bulk
                element.send_keys(text)
                break
    except Exception as e:
        logger.warning(f"Error in human typing: {e}")
        try:
            element.send_keys(text)
        except: pass


def copy_image_to_clipboard(image_path):
    """Copy image file to clipboard using PowerShell"""
    try:
        abs_path = os.path.abspath(image_path)
        cmd = f'''
        Add-Type -AssemblyName System.Windows.Forms
        $files = [System.Collections.Specialized.StringCollection]::new()
        $files.Add("{abs_path}")
        [System.Windows.Forms.Clipboard]::SetFileDropList($files)
        '''
        subprocess.run(["powershell", "-command", cmd], timeout=5)
        return True
    except Exception as e:
        logger.error(f"Failed to copy to clipboard: {e}")
        return False


def upload_image_to_editor(driver, file_path: str) -> bool:
    """LiveJournal editor me image upload karne ki koshish kare (Clipboard + File Input)"""
    try:
        if not os.path.exists(file_path):
            logger.error(f"❌ Image file not found: {file_path}")
            return False

        logger.info(f"🖼️ Uploading image: {file_path}")

        # Method 1: Clipboard Paste (Best for rich text editors like LiveJournal)
        try:
            if copy_image_to_clipboard(file_path):
                time.sleep(1)
                # Paste (Ctrl+V)
                actions = ActionChains(driver)
                actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL)
                actions.perform()
                logger.info("✅ Pasted image from clipboard")
                time.sleep(5) # Wait for upload
                return True
        except Exception as e:
            logger.warning(f"Clipboard paste failed: {e}")

        # Method 2: File Input Fallback
        # Try to find existing file input (LiveJournal usually has one for post editor)
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        target_input = None
        
        # Filter visible or likely inputs
        if inputs:
            target_input = inputs[0]
            
        # If no input found, try to trigger it via "Insert Image" button
        if not target_input:
            img_btns = [
                "button[aria-label*='Image']",
                "button[title*='Image']",
                "button[aria-label*='Photo']",
                "span[class*='svg-icon-image']",
                "//*[contains(@class, 'b-icon-add-image')]",
            ]
            for sel in img_btns:
                try:
                    els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                    for el in els:
                        if el.is_displayed():
                            logger.info(f"Clicking image button: {sel}")
                            driver.execute_script("arguments[0].click();", el)
                            time.sleep(1)
                            inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
                            if inputs:
                                target_input = inputs[0]
                                break
                    if target_input: break
                except: pass

        if target_input:
            try:
                # Unhide if necessary
                driver.execute_script("arguments[0].style.display = 'block'; arguments[0].style.visibility = 'visible';", target_input)
                target_input.send_keys(file_path)
                logger.info("✅ Image path sent to file input")
                time.sleep(8) # Wait for upload (images can take time)
                return True
            except Exception as e:
                logger.error(f"Failed to send keys to input: {e}")
        else:
            logger.warning("⚠️ Could not find file input for image upload")
        
        return False

    except Exception as e:
        logger.error(f"Image upload error: {e}")
        return False


def _resolve_image_path(path_str: str) -> str:
    try:
        base_dir = os.path.dirname(os.path.abspath(BLOG_FILE)) if BLOG_FILE else os.getcwd()
        s = (path_str or "").strip()
        if not s:
            return None
        candidates = []
        if ('/' in s) or ('\\' in s):
            candidates.extend([
                s,
                os.path.abspath(s) if os.path.isabs(s) else os.path.abspath(os.path.join(base_dir, s)),
                os.path.join(base_dir, s.replace('../', '').replace('..\\', '')),
                s.replace('../', '').replace('..\\', ''),
            ])
        else:
            candidates.extend([
                os.path.abspath(os.path.join(base_dir, '..', 'Image', s)),
                os.path.join(base_dir, 'Image', s),
                os.path.join(base_dir, s),
                os.path.join(base_dir, 'Ima', s),
                os.path.abspath(os.path.join(base_dir, '..', 'Ima', s)),
            ])
        for p in candidates:
            try:
                ap = os.path.abspath(p)
                if os.path.isfile(ap):
                    return ap
            except Exception:
                continue
        return None
    except Exception:
        return None


def create_post(driver, title: str, content: str, username: str = "") -> bool:
    try:
        logger.info("📝 LiveJournal blog post create कर रहा हूं...")
        
        urls = [
            "https://www.livejournal.com/post/?draft=https://kavikr543.livejournal.com/d2.html",
            "https://www.livejournal.com/post/",
            "https://www.livejournal.com/entry/new/",
            "https://www.livejournal.com/editjournal.bml",
        ]
        
        ok = False
        for u in urls:
            try:
                driver.get(u)
                time.sleep(5)
                ok = True
                break
            except Exception:
                continue
        
        if not ok:
            logger.error("❌ Could not access post creation page")
            return False
        
        title_el = None
        t_sels = [
            "input[name*='subject']",
            "input[id*='subject']",
            "input[placeholder*='Title']",
            "textarea[name*='subject']",
            "h1[contenteditable='true']",
            "div[contenteditable='true'][data-placeholder*='Title']",
            "*[contenteditable='true'][aria-label*='Title']",
            "//input[contains(@name,'subject') or contains(@id,'subject') or contains(@placeholder,'Title')]",
            "//*[(@contenteditable='true') and (contains(@data-placeholder,'Title') or contains(@aria-label,'Title'))]",
        ]
        
        for sel in t_sels:
            try:
                els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if True:
                        title_el = el
                        logger.info(f"✅ Title field found: {sel}")
                        break
                if title_el:
                    break
            except Exception:
                continue

        if not title_el:
            try:
                js = """
                return document.querySelector('[data-placeholder*="Title"],[aria-label*="Title"],h1[contenteditable="true"],input[placeholder*="Title"],textarea[placeholder*="Title"]');
                """
                title_el = driver.execute_script(js)
            except Exception:
                title_el = None
        if not (title_el):
            try:
                js_list = """
                const nodes = Array.from(document.querySelectorAll('[contenteditable="true"]'));
                const scored = nodes.map(n=>{
                  const rect = n.getBoundingClientRect();
                  const ph = (n.getAttribute('data-placeholder')||'').toLowerCase();
                  const aria = (n.getAttribute('aria-label')||'').toLowerCase();
                  return {el:n, top:rect.top, ph, aria};
                }).sort((a,b)=>a.top-b.top);
                return scored.map(s=>s.el);
                """
                nodes = driver.execute_script(js_list)
                if nodes and len(nodes) >= 1:
                    title_el = nodes[0]
                    logger.info("✅ Title element chosen from contenteditable list (topmost)")
                # Also set editor from list if not found
                if not 'editor_el' in locals() or not editor_el:
                    if nodes and len(nodes) >= 2:
                        editor_el = nodes[1]
                        logger.info("✅ Content editor chosen from contenteditable list (second)")
            except Exception:
                pass
        
        editor_el = None
        e_sels = [
            "div[role='textbox'][contenteditable='true']",
            "div[contenteditable='true'][data-placeholder*='Start typing']",
            "div[contenteditable='true']",
            "textarea[name*='event']",
            "textarea[id*='event']",
            "//div[@role='textbox' and @contenteditable='true']",
            "//textarea[contains(@name,'event') or contains(@id,'event')]",
            "//textarea",
        ]
        
        for sel in e_sels:
            try:
                els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed():
                        editor_el = el
                        logger.info(f"✅ Content editor found: {sel}")
                        break
                if editor_el:
                    break
            except Exception:
                continue
        
        def _enter_title_text(drv, el, txt):
            try:
                el.click()
                time.sleep(0.2)
            except Exception:
                pass
            try:
                el.clear()
            except Exception:
                pass
            try:
                el.send_keys(Keys.CONTROL + "a")
                el.send_keys(Keys.DELETE)
            except Exception:
                pass
            try:
                human_type(el, txt) # Use slow typing
                return True
            except Exception:
                pass
            try:
                drv.execute_script("arguments[0].focus();", el)
                drv.execute_script("arguments[0].innerText = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el, txt)
                return True
            except Exception:
                return False

        def _enter_text(drv, el, txt):
            return _enter_title_text(drv, el, txt)

        def _enter_body_text(drv, el, txt):
            try:
                logger.info(f"📝 Entering body text ({len(txt)} chars) with link and image support")
                el.click()
                time.sleep(0.5)
                
                # Content को साफ करने के लिए
                try:
                    el.send_keys(Keys.CONTROL + "a")
                    el.send_keys(Keys.BACKSPACE)
                except:
                    pass
                
                # Regex for [text](url) OR {image_path}
                pattern = r'(\[([^\]]+)\]\(([^)]+)\))|(\{([^{}]+)\})'
                
                last_end = 0
                for match in re.finditer(pattern, txt):
                    plain_text = txt[last_end:match.start()]
                    if plain_text:
                        human_type(el, plain_text) # Use slow typing
                        time.sleep(0.1)
                    new_last_end = match.end()
                    
                    if match.group(1): # Link
                        link_text = match.group(2)
                        link_url = match.group(3)
                        
                        logger.info(f"🔗 Creating backlink: '{link_text}' -> {link_url}")
                        try:
                            human_type(el, link_text) # Use slow typing
                            time.sleep(0.5)
                            
                            actions = ActionChains(drv)
                            actions.key_down(Keys.SHIFT)
                            for _ in range(len(link_text)):
                                actions.send_keys(Keys.LEFT)
                            actions.key_up(Keys.SHIFT)
                            actions.perform()
                            time.sleep(0.5)
                            
                            actions = ActionChains(drv)
                            actions.key_down(Keys.CONTROL)
                            actions.send_keys('k')
                            actions.key_up(Keys.CONTROL)
                            actions.perform()
                            time.sleep(1.5)
                            
                            actions = ActionChains(drv)
                            actions.send_keys(link_url)
                            time.sleep(0.5)
                            actions.send_keys(Keys.ENTER)
                            actions.perform()
                            time.sleep(1.0)
                            
                            drv.execute_script("""
                                var url = arguments[0];
                                if (!document.queryCommandValue('createLink')) {
                                    document.execCommand('createLink', false, url);
                                }
                            """, link_url)
                            
                            el.send_keys(Keys.RIGHT)
                            el.send_keys(" ")
                            
                        except Exception as e:
                            logger.error(f"Link insertion error: {e}")
                            el.send_keys(f" {link_url} ")
                        
                        last_end = match.end()

                    elif match.group(4): # Image
                        image_path = match.group(5)
                        logger.info(f"🖼️ Found image placeholder: {image_path}")
                        current_match_end = match.end()
                        
                        resolved = _resolve_image_path(image_path)
                        if resolved and upload_image_to_editor(drv, resolved):
                            logger.info("✅ Image uploaded")
                            try:
                                time.sleep(2)
                                actions = ActionChains(drv)
                                # Ensure cursor is after image
                                actions.send_keys(Keys.RIGHT)
                                actions.pause(0.5)
                                # Single Enter to move to next line
                                actions.send_keys(Keys.ENTER)
                                actions.perform()
                                time.sleep(0.5)
                                
                                # Skip any following newlines in the source text
                                try:
                                    i = match.end()
                                    while i < len(txt) and txt[i] in '\r\n':
                                        i += 1
                                    current_match_end = i
                                except:
                                    pass
                            except: pass
                        else:
                            logger.error("❌ Image upload failed")
                            try:
                                human_type(el, "{" + image_path + "}")
                            except:
                                pass
                        
                        last_end = current_match_end
                
                remaining_text = txt[last_end:]
                if remaining_text:
                    human_type(el, remaining_text) # Use slow typing
                
                logger.info("✅ Body text with hyperlinks/images entered successfully")
                return True
            except Exception as e:
                logger.error(f"Body entry error: {e}")
                return False

        # Enter title if found
        if title_el:
            try:
                _enter_title_text(driver, title_el, title)
                logger.info(f"✅ Title entered: '{title}'")
            except Exception as e:
                logger.error(f"Title entry error: {e}")
        
        # Enter content if found
        if editor_el:
            try:
                _enter_body_text(driver, editor_el, content)
                time.sleep(1)
            except Exception as e:
                logger.error(f"Content entry error: {e}")

        try:
            filled_ok = False
            title_len = 0
            body_len = 0
            if title_el:
                try:
                    title_len = int(driver.execute_script("var e=arguments[0];return (e.value||e.textContent||'').trim().length;", title_el))
                except Exception:
                    title_len = 0
            if editor_el:
                try:
                    body_len = int(driver.execute_script("var e=arguments[0];return (e.value||e.innerText||e.textContent||'').trim().length;", editor_el))
                except Exception:
                    body_len = 0
            filled_ok = (body_len > 0) and (title_len > 0 or len(title.strip()) > 0)
            if not filled_ok and title_el:
                _enter_text(driver, title_el, title if title.strip() else "Untitled")
                time.sleep(0.5)
        except Exception:
            pass
        
        b_sels = [
            "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'tune in and publish')]",
            "button[data-action*='publish']",
            "*[role='button'][data-qa-id*='publish']",
            "button[type='submit']",
        ]
        
        for sel in b_sels:
            try:
                els = driver.find_elements(By.XPATH, sel) if sel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        logger.info(f"✅ Post button found: {sel}")
                        logger.info(f"   Button text: '{el.text or el.get_attribute('value')}'")
                        
                        # Scroll into view
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.5)
                        
                        # Highlight for debugging
                        driver.execute_script("arguments[0].style.border='3px solid green';", el)
                        time.sleep(0.5)
                        
                        try:
                            el.click()
                            logger.info("✅ Publish menu opened")
                        except Exception:
                            driver.execute_script("arguments[0].click();", el)
                        
                        # Remove highlight
                        driver.execute_script("arguments[0].style.border='';", el)
                        
                        time.sleep(2)
                        
                        p_sels = [
                            "//button[contains(.,'Publish')]",
                            "//a[contains(.,'Publish')]",
                            "button[data-action*='confirm-publish']",
                        ]
                        
                        for psel in p_sels:
                            try:
                                pels = driver.find_elements(By.XPATH, psel) if psel.startswith('//') else driver.find_elements(By.CSS_SELECTOR, psel)
                                for pel in pels:
                                    if pel.is_displayed() and pel.is_enabled():
                                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", pel)
                                        time.sleep(0.5)
                                        try:
                                            pel.click()
                                        except Exception:
                                            driver.execute_script("arguments[0].click();", pel)
                                        time.sleep(5)
                                        current_url = driver.current_url.lower()
                                        page_source = driver.page_source.lower()
                                        if ('can\'t publish empty blog' in page_source) or ('empty blog' in page_source):
                                            try:
                                                if title_el:
                                                    _enter_title_text(driver, title_el, title if title.strip() else "Untitled")
                                                if editor_el:
                                                    _enter_body_text(driver, editor_el, content)
                                                time.sleep(1)
                                                try:
                                                    el.click()
                                                except Exception:
                                                    driver.execute_script("arguments[0].click();", el)
                                                time.sleep(1)
                                                for p2 in p_sels:
                                                    try:
                                                        pp = driver.find_elements(By.XPATH, p2) if p2.startswith('//') else driver.find_elements(By.CSS_SELECTOR, p2)
                                                        for q in pp:
                                                            if q.is_displayed() and q.is_enabled():
                                                                try:
                                                                    q.click()
                                                                except Exception:
                                                                    driver.execute_script("arguments[0].click();", q)
                                                                time.sleep(5)
                                                                current_url = driver.current_url.lower()
                                                                page_source = driver.page_source.lower()
                                                                break
                                                    except Exception:
                                                        continue
                                            except Exception:
                                                pass
                                        if ('draft saved' in page_source):
                                            logger.info("⚠️ Still shows 'Draft saved' — verifying on journal page")
                                        # Verify by visiting journal home
                                        if username:
                                            try:
                                                driver.get(f"https://{username}.livejournal.com/")
                                                time.sleep(5)
                                                src = driver.page_source.lower()
                                                if title.lower() in src or content[:50].lower() in src:
                                                    logger.info("✅ Blog visible on journal homepage")
                                                    return True
                                            except Exception:
                                                pass
                                        # Fallback heuristic
                                        if ('posted' in page_source) or ('entry' in current_url):
                                            logger.info("✅ Blog published (heuristic)")
                                            return True
                            except Exception:
                                continue
                        
                        return True
            except Exception:
                continue
        
        logger.error("❌ Could not find post button")
        return False
        
    except Exception as e:
        logger.error(f"❌ LiveJournal post creation error: {e}")
        return False

 
def run(driver, website: str, email: str, username: str, password: str) -> bool:
    try:
        user = (username or email or "").strip()
        if not user or not password:
            logger.error("Missing credentials for LiveJournal")
            return False
        target = website or "https://www.livejournal.com/"
        if "livejournal" not in target.lower():
            target = "https://www.livejournal.com/"
        try:
            driver.get(target)
            time.sleep(4)
        except Exception as e:
            logger.error(f"Navigation error: {e}")
            return False
        try:
            click_login_button(driver)
        except Exception:
            pass
        ok = fill_login_form(driver, user, password)
        time.sleep(3)
        logged = is_logged_in(driver)
        blog = read_blog_text()
        title = blog.get("title", "") or "Automated LiveJournal Post"
        content = blog.get("body", "") or ""
        posted = False
        if ok and logged:
            posted = create_post(driver, title, content, user)
        return bool(ok and logged and posted)
    except Exception as e:
        logger.error(f"LiveJournal run error: {e}")
        return False
 
def main():
    install_packages()
    creds = load_livejournal_credentials()
    if not creds or not creds.get('password') or not (creds.get('username') or creds.get('email')):
        print("credentials.xlsx me livejournal.com ke liye valid credentials nahi mile")
        return
    
    username = creds.get('username') or creds.get('email')
    password = creds.get('password')
    
    print(f"🔐 Using credentials for LiveJournal:")
    print(f"   Username/Email: {username}")
    print(f"   Password: {'*' * len(password)}")
    
    driver = create_driver()
    if not driver:
        print("Chrome driver create nahi ho paya")
        return
    
    try:
        # Step 1: Go to LiveJournal homepage
        print("\n🌐 LiveJournal homepage par ja raha hoon...")
        driver.get("https://www.livejournal.com/")
        time.sleep(4)
        
        screenshot(driver, "livejournal_homepage")
        
        # Step 2: Click login button on homepage (image 1 में जो दिख रहा है)
        print("🔘 Homepage par login button dhoondh raha hoon...")
        clicked = click_login_button(driver)
        
        if not clicked:
            print("❌ Login button nahi mila")
        
        # Step 3: Fill login form
        print("📝 Login form fill kar raha hoon...")
        ok = fill_login_form(driver, username, password)
        time.sleep(4)
        
        # Step 4: Check if login successful
        logged = is_logged_in(driver)
        screenshot(driver, "after_login")
        
        if not (ok and logged):
            print("❌ LiveJournal login fail")
            return
        
        print("✅ LiveJournal login successful!")
        
        # Step 5: Create blog post
        title = "Automated Blog Post"
        content = "Default content"
        
        if os.path.exists("blog.txt"):
            try:
                with open("blog.txt", "r", encoding="utf-8") as f:
                    lines = f.readlines()
                
                if lines:
                    title = lines[0].strip()
                    content = "".join(lines[1:]).strip()
                    print(f"📖 Read blog.txt - Title: {title}")
                else:
                    print("⚠️ blog.txt is empty")
            except Exception as e:
                print(f"Error reading blog.txt: {e}")
        else:
            print("⚠️ blog.txt not found")
        
        print(f"📝 Blog post create kar raha hoon...")
        print(f"   Title: {title}")
        print(f"   Content length: {len(content)} characters")
        
        posted = create_post(driver, title, content, username)
        time.sleep(2)
        screenshot(driver, "after_submit")
        
        if posted:
            print("✅✅✅ LiveJournal par blog post successful! ✅✅✅")
        else:
            print("❌ LiveJournal par blog post karne me problem aayi")
            
    finally:
        # Keep browser open for a while to see results
        print("\n⏸️ Browser 10 seconds ke liye open rahega...")
        time.sleep(10)
        
        try:
            if hasattr(driver, 'service') and driver.service:
                driver.service.stop()
        except Exception:
            pass
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == '__main__':
    main()
